"""손익분석 - 전처리 모듈

업로드된 파일들을 DataFrame 으로 변환하고 상품ID 통합까지 책임진다.

구성:
    1. 상수 정의
    2. 공통 헬퍼 (DataFrame / lookup / 엑셀 I/O)
    3. 전처리 - 기초 DB
    4. 차량 상세 / lookup 빌더
    5. 상품ID 모음 (collect_product_ids)
    6. 전처리 - 매입원가 (상품원장 / 폐자원 / 페이백 / 일반)
    7. 전처리 - 제조원가 (재료비 / 노무비·경비 / 직접경비)
"""

from io import BytesIO
import functools
import re

import numpy as np
import pandas as pd


# ============================================================
# 1. 상수 정의
# ============================================================

DETAIL_COLUMNS = [
    "신번호", "구번호", "차대번호", "차종", "차명",
    "반납일자", "매입일자", "분류1", "분류2", "분류3", "분류4",
]
OUTPUT_DETAIL_COLUMNS = [*DETAIL_COLUMNS, "매입연도", "매입월"]

PRODUCT_ID_SOURCE_KEYS = [
    "매입조회", "검사매출", "정비매출", "기초재고",
    "전체상품조회", "위탁수불부_기초", "위탁수불부_입고",
]
PRODUCT_ID_COLUMNS = [
    "회계연도", "회계월", "상품ID",
    "기초재고", "정상입고", "타처입고",
    "당사/타사", "매출구분", "선매입여부",
    *OUTPUT_DETAIL_COLUMNS,
]

# 금액 컬럼
PURCHASE_AMOUNT_COLUMNS = ["상품매입액", "취득세", "매입수수료"]
WASTE_RESOURCE_COLUMN = "폐자원공제"
PAYBACK_RETURN_COLUMN = "페이백(반납)"
PAYBACK_UNRETURNED_COLUMN = "페이백(미반납)"
EXCESS_DRIVING_COLUMN = "초과운행"
DIFFERENCE_ALLOCATION_COLUMN = "차액배부"
TOTAL_PURCHASE_COST_COLUMN = "매입원가"  # FINAL_COST_MONTHLY_COLUMNS 8개 합계용

FINAL_COST_AMOUNT_COLUMNS = [
    *PURCHASE_AMOUNT_COLUMNS,
    WASTE_RESOURCE_COLUMN,
    PAYBACK_RETURN_COLUMN,
    PAYBACK_UNRETURNED_COLUMN,
    EXCESS_DRIVING_COLUMN,
]
FINAL_COST_MONTHLY_COLUMNS = [
    *PURCHASE_AMOUNT_COLUMNS,
    WASTE_RESOURCE_COLUMN,
    DIFFERENCE_ALLOCATION_COLUMN,
    PAYBACK_RETURN_COLUMN,
    PAYBACK_UNRETURNED_COLUMN,
    EXCESS_DRIVING_COLUMN,
]
DIFFERENCE_SOURCE_COST_COLUMNS = ["취득세", "매입수수료", WASTE_RESOURCE_COLUMN]
PRODUCT_LEDGER_TOTAL_COST_COLUMNS = [
    *DIFFERENCE_SOURCE_COST_COLUMNS,
    PAYBACK_RETURN_COLUMN,
    PAYBACK_UNRETURNED_COLUMN,
    EXCESS_DRIVING_COLUMN,
]

# 시트 컬럼 후보 (시트마다 표기가 다른 경우)
WASTE_RESOURCE_AMOUNT_COLUMNS = ["매입세금공제액", "매입세액공제액", "매입세액공제"]
PAYBACK_RETURN_AMOUNT_COLUMNS = ["선수수익(VAT外)", "선수수익(VAT외)", "선수수익(VAT 外)"]

MATERIAL_SALES_COLUMNS = ["정비매출", "사내매출", "위탁매출"]

# 재료비 시트 금액 컬럼 후보 (시트마다 표기가 다를 수 있음)
MATERIAL_COST_AMOUNT_COLUMNS = ["금액", "출고금액", "출고가액", "원가", "재료비"]

# 기초 DB DataFrame 키
BASE_DF_KEYS = [
    "매입조회", "검사매출", "정비매출",
    "기초재고", "기초재고_전체",
    "전체상품조회",
    "위탁수불부", "위탁수불부_전체", "위탁수불부_기초", "위탁수불부_입고",
]
HIDDEN_BASE_DF_KEYS = ["기초재고_전체", "위탁수불부_전체"]


def _sort_product_id_df(df):
    """1번 원가대상 생성 직후 적용할 행 정렬."""
    if df is None or df.empty:
        return df

    work = df.copy()
    sales_order = {"사내매출": 0, "위탁매출": 1, "검사매출": 2, "정비매출": 3}
    sales_key = (
        work["매출구분"].astype(str).str.strip().map(sales_order).fillna(99)
        if "매출구분" in work.columns
        else pd.Series(99, index=work.index)
    )

    def _num_col(column):
        if column not in work.columns:
            return pd.Series(0, index=work.index)
        return pd.to_numeric(work[column], errors="coerce").fillna(0)

    qty_key = pd.Series(3, index=work.index)
    qty_key = qty_key.mask(_num_col("타처입고").eq(1), 2)
    qty_key = qty_key.mask(_num_col("정상입고").eq(1), 1)
    qty_key = qty_key.mask(_num_col("기초재고").eq(1), 0)

    own_mask = (
        work["당사/타사"].astype(str).str.strip().eq("당사차량")
        if "당사/타사" in work.columns
        else pd.Series(False, index=work.index)
    )
    purchase_date = (
        pd.to_datetime(work["매입일자"], errors="coerce")
        if "매입일자" in work.columns
        else pd.Series(pd.NaT, index=work.index)
    )
    return_date = (
        pd.to_datetime(work["반납일자"], errors="coerce")
        if "반납일자" in work.columns
        else pd.Series(pd.NaT, index=work.index)
    )
    date_key = purchase_date.where(own_mask, return_date)

    work["_sales_key"] = sales_key.values
    work["_qty_key"] = qty_key.values
    work["_date_key"] = date_key.values
    work["_product_id_key"] = (
        work["상품ID"].astype(str).str.strip()
        if "상품ID" in work.columns
        else pd.Series("", index=work.index)
    ).values

    return (
        work.sort_values(
            by=["_sales_key", "_qty_key", "_date_key", "_product_id_key"],
            ascending=[True, True, True, True],
            na_position="last",
            kind="stable",
        )
        .drop(columns=["_sales_key", "_qty_key", "_date_key", "_product_id_key"])
        .reset_index(drop=True)
    )


# ============================================================
# 2. 공통 헬퍼
# ============================================================

# ----- DataFrame 기본 처리 -----

def _strip_columns(df):
    """컬럼명 좌우 공백 제거."""
    df = df.copy()
    df.columns = [str(column).strip() for column in df.columns]
    return df


def _is_flag_one(series):
    """숫자 1 또는 문자열 '1'인지 확인."""
    numeric_values = pd.to_numeric(series, errors="coerce")
    text_values = series.astype(str).str.strip()
    return numeric_values.eq(1) | text_values.eq("1")


def _ensure_product_id(df):
    """상품ID 컬럼이 없으면 차량아이디 / CODE 에서 생성."""
    df = df.copy()
    if "상품ID" not in df.columns and "차량아이디" in df.columns:
        df["상품ID"] = df["차량아이디"]
    if "상품ID" not in df.columns and "CODE" in df.columns:
        df["상품ID"] = df["CODE"]
    return df


def _get_first_existing_column(df, candidates):
    """후보 컬럼 중 가장 먼저 존재하는 컬럼 반환. 없으면 빈 문자열."""
    for column in candidates:
        if column in df.columns:
            return df[column]
    return ""


def _normalize_lookup_value(value):
    """lookup 키로 사용하기 위한 정규화 (공백 제거)."""
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def _set_accounting_year_month(df, date_column, year_col="회계연도", month_col="회계월"):
    """날짜 컬럼에서 회계연도/회계월(Int64) 컬럼 생성."""
    if date_column in df.columns:
        parsed = pd.to_datetime(df[date_column], format="mixed", errors="coerce")
        df[year_col] = parsed.dt.year.astype("Int64")
        df[month_col] = parsed.dt.month.astype("Int64")
    else:
        df[year_col] = pd.NA
        df[month_col] = pd.NA
    return df


def _load_excel_sheets(file, clean=True):
    """엑셀 파일에서 모든 시트를 dict로 로드. clean=True 면 컬럼 strip + 빈 row/col 제거."""
    file.seek(0)
    sheets = pd.read_excel(file, sheet_name=None)
    if not clean:
        return sheets

    processed = {}
    for sheet_name, df in sheets.items():
        df = _strip_columns(df)
        df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
        processed[sheet_name] = df
    return processed


# ----- 상품ID 기반 머지/할당 -----

def _merge_by_product_id(target_df, source_df, value_columns):
    """상품ID 정규화 키로 source_df의 value_columns 를 target_df에 left merge."""
    target = target_df.copy()
    target["_상품ID_lookup"] = target["상품ID"].astype(str).str.strip()

    source = source_df.copy()
    source["_상품ID_lookup"] = source["상품ID"].astype(str).str.strip()
    source = source[["_상품ID_lookup", *value_columns]]

    merged = target.merge(source, on="_상품ID_lookup", how="left")
    return merged.drop(columns=["_상품ID_lookup"])


def _excel_round(value, digits=0):
    """Excel ROUND 와 같은 반올림. 0.5 는 0에서 먼 방향으로 반올림한다."""
    from decimal import Decimal, ROUND_HALF_UP

    if pd.isna(value):
        return np.nan

    try:
        digits = int(digits)
        # str() 변환으로 부동소수점을 십진수 표현으로 바꾼 뒤 Decimal로 정확하게 처리
        # epsilon 보정은 오히려 x.4999...를 올림하는 과보정이 생기므로 제거
        number = Decimal(str(value))
        quantizer = Decimal("1").scaleb(-digits)
        rounded = number.quantize(quantizer, rounding=ROUND_HALF_UP)
    except Exception:
        return value

    if digits == 0:
        return int(rounded)
    return float(rounded)


def _excel_round_series(series, digits=0):
    """Series/DataFrame 컬럼에 Excel ROUND 방식 적용."""
    numeric = pd.to_numeric(series, errors="coerce")
    # str() 변환 후 Decimal로 처리하는 _excel_round를 직접 적용
    # NaN이 아닌 값만 처리하고 나머지는 NaN 유지
    result = numeric.copy()
    mask = numeric.notna()
    if mask.any():
        result[mask] = numeric[mask].apply(lambda v: _excel_round(v, digits))
    return result


def _allocate_amount(final_df, target_column, total_amount, row_mask):
    """row_mask 행에 total_amount 를 균등 배부 (잔여는 첫 행에 가산)."""
    if total_amount == 0:
        return
    allocation_count = int(row_mask.sum())
    if allocation_count == 0:
        return

    allocation_value = _excel_round(total_amount / allocation_count)
    final_df.loc[row_mask, target_column] = allocation_value

    remainder = total_amount - allocation_value * allocation_count
    if remainder != 0:
        first_index = final_df.index[row_mask][0]
        final_df.loc[first_index, target_column] += remainder


def _normal_inbound_mask(final_df):
    """정상입고==1 행 마스크."""
    return pd.to_numeric(final_df["정상입고"], errors="coerce").fillna(0).eq(1)


# ----- 엑셀 출력 -----

# 컬럼명에 이 키워드가 포함되면 숫자 서식(#,##0 등)을 적용하지 않음 (연도에 천단위 콤마 방지)
_YEAR_COLUMN_KEYWORDS = ("연도", "년도")


def _is_year_column(column_name):
    return any(kw in str(column_name) for kw in _YEAR_COLUMN_KEYWORDS)


def _apply_year_column_plain_format(worksheet, df):
    """연도/년도 가 포함된 컬럼은 천단위 콤마 없는 일반 숫자 서식으로 고정."""
    for col_idx, column_name in enumerate(df.columns, start=1):
        if not _is_year_column(column_name):
            continue
        for row_offset in range(len(df)):
            cell = worksheet.cell(row=2 + row_offset, column=col_idx)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = "0"


def dataframe_to_excel_bytes(df, sheet_name="Sheet1"):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        _apply_excel_column_group_styles(writer.sheets[sheet_name], df)
        _apply_year_column_plain_format(writer.sheets[sheet_name], df)
    output.seek(0)
    return output.getvalue()


# ===== 최종원가 컬럼 그룹별 색상 =====
# 그룹 A: 수량/금액/출고/기말 + 원가동인/유효실측시간
_COLUMN_GROUP_A = [
    "기초_수량", "기초_금액", "정상입고_수량", "정상입고_금액",
    "타처입고_수량", "타처입고_금액", "제조원가",
    "정상출고_수량", "정상출고_금액", "자산출고_수량", "자산출고_금액",
    "기타출고_수량", "기타출고_금액", "기말_수량", "기말_금액",
    "RTC_일수", "SM_일수",
    "유효실측시간_전체", "유효실측시간_RQI", "유효실측시간_정비",
    "유효실측시간_판금", "유효실측시간_도장",
]
# 그룹 B: 매출원가/매입원가/제조원가/항목별/페이백/재료비·노무비·제조경비 누적
_COLUMN_GROUP_B = [
    "매출원가_누적합계", "매출원가_전월누적", "매출원가_당월",
    "매입원가_누적합계", "매입원가_전월누적", "매입원가_당월",
    "제조원가_누적합계", "제조원가_전월누적", "제조원가_당월",
    "상품매입액_합계", "상품매입액_전월", "상품매입액_당월",
    "취득세_합계", "취득세_전월", "취득세_당월",
    "매입수수료_합계", "매입수수료_전월", "매입수수료_당월",
    "폐자원공제_합계", "폐자원공제_전월", "폐자원공제_당월",
    "차액배부_합계", "차액배부_전월", "차액배부_당월",
    "페이백_합계", "페이백_전월", "페이백_당월",
    "페이백(반납)_합계", "페이백(반납)_전월", "페이백(반납)_당월",
    "페이백(미반납)_합계", "페이백(미반납)_전월", "페이백(미반납)_당월",
    "초과운행_합계", "초과운행_전월", "초과운행_당월",
    "재료비_누적합계", "재료비_전월누적", "재료비_당월",
    "노무비_누적합계", "노무비_전월누적", "노무비_당월",
    "제조경비_누적합계", "제조경비_전월누적", "제조경비_당월",
]
_GROUP_A_COLOR = "DDEBF7"   # 옅은 파랑
_GROUP_B_COLOR = "E2EFDA"   # 옅은 초록


def _column_group_of(column_name):
    """컬럼명이 속한 그룹 반환: 'A' / 'B' / None."""
    name = str(column_name)
    if name in _COLUMN_GROUP_A:
        return "A"
    if name in _COLUMN_GROUP_B:
        return "B"
    return None


def _apply_excel_column_group_styles(worksheet, df):
    """엑셀 시트에 컬럼 그룹별 배경색 + 헤더 볼드/큰글씨 적용."""
    try:
        from openpyxl.styles import PatternFill, Font
    except Exception:
        return

    fill_a = PatternFill(start_color=_GROUP_A_COLOR, end_color=_GROUP_A_COLOR, fill_type="solid")
    fill_b = PatternFill(start_color=_GROUP_B_COLOR, end_color=_GROUP_B_COLOR, fill_type="solid")
    n_rows = len(df)

    for col_idx, column_name in enumerate(df.columns, start=1):
        group = _column_group_of(column_name)
        # 헤더(1행): 볼드 + 글씨 크게 (그룹이면 색도)
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = Font(bold=True, size=12)
        if group is None:
            continue
        fill = fill_a if group == "A" else fill_b
        header_cell.fill = fill
        # 데이터 셀(2행~)에 색칠
        for row_offset in range(n_rows):
            worksheet.cell(row=2 + row_offset, column=col_idx).fill = fill


def _safe_excel_sheet_name(name, used_names):
    """엑셀 시트명 규칙(31자, 특수문자 등) 적용 + 중복 회피."""
    invalid_chars = ["\\", "/", "*", "?", ":", "[", "]"]
    safe_name = str(name).strip() or "Sheet"
    for char in invalid_chars:
        safe_name = safe_name.replace(char, "_")
    safe_name = safe_name[:31] or "Sheet"

    base_name = safe_name
    index = 1
    while safe_name in used_names:
        index += 1
        suffix = f"_{index}"
        safe_name = f"{base_name[:31 - len(suffix)]}{suffix}"

    used_names.add(safe_name)
    return safe_name


def workbook_to_excel_bytes(sheet_dfs):
    output = BytesIO()
    used_sheet_names = set()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheet_dfs.items():
            safe_sheet_name = _safe_excel_sheet_name(sheet_name, used_sheet_names)
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name)
            _apply_year_column_plain_format(writer.sheets[safe_sheet_name], df)
    output.seek(0)
    return output.getvalue()


def styled_dataframe_for_display(df):
    """화면 표시용 Styler: 컬럼 그룹별 배경색 + 헤더 볼드/큰글씨.

    st.dataframe(styled_dataframe_for_display(df)) 형태로 사용.
    그룹 컬럼이 하나도 없으면 일반 display DataFrame 을 그대로 반환.
    셀 수가 많으면 Styler 렌더 한도를 자동으로 올린다.
    """
    display_df = dataframe_for_display(df)

    group_map = {col: _column_group_of(col) for col in display_df.columns}
    if not any(group_map.values()):
        return display_df  # 색칠 대상 없음

    # Styler 렌더 셀 수 한도 자동 상향 (기본 262,144)
    n_cells = display_df.shape[0] * display_df.shape[1]
    try:
        current_max = pd.get_option("styler.render.max_elements")
        if n_cells > current_max:
            pd.set_option("styler.render.max_elements", int(n_cells * 1.1) + 1)
    except Exception:
        pass

    color_map = {"A": f"#{_GROUP_A_COLOR}", "B": f"#{_GROUP_B_COLOR}"}

    def _col_style(col):
        g = group_map.get(col.name)
        if g is None:
            return ["" for _ in col]
        return [f"background-color: {color_map[g]};" for _ in col]

    def _header_style():
        styles = []
        for col in display_df.columns:
            g = group_map.get(col)
            bg = f"background-color: {color_map[g]}; " if g else ""
            styles.append({
                "selector": f"th.col_heading.col{list(display_df.columns).index(col)}",
                "props": f"{bg}font-weight: bold; font-size: 13px;",
            })
        return styles

    styler = display_df.style.apply(_col_style, axis=0)
    styler = styler.set_table_styles(_header_style(), overwrite=False)
    return styler


def dataframe_for_display(df):
    """Streamlit dataframe 표시용 (object/string 컬럼의 NaN → '', 날짜 컬럼 시간 제거)."""
    display_df = df.copy()
    display_df.columns = [str(column) for column in display_df.columns]
    for column in display_df.columns:
        # 연도/년도 컬럼: 숫자형이면 문자열로 바꿔 화면에서 천단위 콤마가 붙지 않게 함
        if _is_year_column(column) and pd.api.types.is_numeric_dtype(display_df[column]):
            display_df[column] = display_df[column].apply(
                lambda v: "" if pd.isna(v) else str(int(v))
            )
            continue
        # datetime 컬럼: 시간이 전부 00:00:00 이면 날짜만(YYYY-MM-DD) 표시
        if pd.api.types.is_datetime64_any_dtype(display_df[column]):
            series = display_df[column]
            non_na = series.dropna()
            times_all_zero = (
                non_na.empty
                or (non_na.dt.hour.eq(0).all()
                    and non_na.dt.minute.eq(0).all()
                    and non_na.dt.second.eq(0).all())
            )
            if times_all_zero:
                display_df[column] = series.dt.strftime("%Y-%m-%d").where(series.notna(), "")
            else:
                display_df[column] = series.dt.strftime("%Y-%m-%d %H:%M:%S").where(series.notna(), "")
            continue
        if (
            pd.api.types.is_object_dtype(display_df[column])
            or pd.api.types.is_string_dtype(display_df[column])
        ):
            display_df[column] = display_df[column].apply(
                lambda value: "" if pd.isna(value) else _strip_zero_time(value)
            )
    return display_df


def _strip_zero_time(value):
    """문자열로 들어온 'YYYY-MM-DD 00:00:00' 형태에서 시간이 0 이면 날짜만 남김."""
    text = str(value)
    # 'YYYY-MM-DD 00:00:00' 패턴
    if text.endswith(" 00:00:00") and len(text) == 19:
        return text[:10]
    # 'YYYY-MM-DD HH:MM:SS' 인데 시간이 00:00:00 인 경우 일반화
    import re as _re
    m = _re.match(r"^(\d{4}-\d{2}-\d{2}) 00:00:00(\.0+)?$", text)
    if m:
        return m.group(1)
    return text


# ============================================================
# 3. 전처리 - 기초 DB
# ============================================================

def preprocess_purchase_inquiry(file):
    """매입조회"""
    df = _strip_columns(pd.read_excel(file))

    if "상품ID" not in df.columns and "차량아이디" in df.columns:
        df["상품ID"] = df["차량아이디"]

    if "회계월" not in df.columns and "계산서일자" in df.columns:
        parsed_month = pd.to_datetime(
            df["계산서일자"], format="mixed", errors="coerce"
        ).dt.month
        if parsed_month.isna().all():
            parsed_month = pd.to_numeric(df["계산서일자"], errors="coerce")
        df["회계월"] = parsed_month

    df["입고구분"] = df["매입유형-분류1"].apply(
        lambda x: "타처입고" if str(x).strip() == "자산" else "정상입고"
    )
    return df


def preprocess_product_master(file):
    """전체상품조회"""
    df = _strip_columns(pd.read_excel(file))
    return _ensure_product_id(df)


def preprocess_consignment_ledger(file, settlement_year, settlement_month):
    """위탁수불부 → (결산연월 전체, 기초==1, 입고==1)."""
    df = _ensure_product_id(_strip_columns(pd.read_excel(file)))

    required_columns = ["회계연도", "회계월", "기초", "입고"]
    missing_columns = [column for column in required_columns if column not in df.columns]
    if missing_columns:
        raise KeyError(
            "위탁수불부 파일에 다음 컬럼이 없습니다: "
            + ", ".join(f"'{column}'" for column in missing_columns)
        )

    year = pd.to_numeric(df["회계연도"], errors="coerce")
    month = pd.to_numeric(df["회계월"], errors="coerce")
    period_mask = year.eq(int(settlement_year)) & month.eq(int(settlement_month))
    period_df = df[period_mask].copy()

    # 출고상태 컬럼 생성: 판매==1 → 위탁판매, 매입==1 → 위탁매입, 취소==1 → 위탁취소
    outbound_conditions = []
    outbound_choices = []
    for col, label in [("판매", "위탁판매"), ("매입", "위탁매입"), ("취소", "위탁취소")]:
        if col in period_df.columns:
            outbound_conditions.append(_is_flag_one(period_df[col]))
            outbound_choices.append(label)
    if outbound_conditions:
        period_df["출고상태"] = np.select(outbound_conditions, outbound_choices, default="")

    opening_df = period_df[_is_flag_one(period_df["기초"])].copy()
    inbound_df = period_df[_is_flag_one(period_df["입고"])].copy()
    return period_df, opening_df, inbound_df


def preprocess_sales(file, exclude_partner=None):
    """검사매출 / 정비매출"""
    df = _ensure_product_id(_strip_columns(pd.read_excel(file)))

    if "세부내역" in df.columns:
        df = df[~df["세부내역"].astype(str).str.contains("보증수리", na=False)].copy()
    if exclude_partner and "거래처" in df.columns:
        partner = df["거래처"].astype(str).str.strip()
        df = df[~partner.eq(str(exclude_partner).strip())].copy()
    if "상품ID" in df.columns:
        df = df.drop_duplicates(subset=["상품ID"]).copy()
    return df


def preprocess_opening_inventory(file, base_df):
    """기초재고 → (전체, 전월 기말여부==1)"""
    df = _ensure_product_id(_strip_columns(pd.read_excel(file)))

    if base_df is None or base_df.empty:
        raise ValueError("기초재고 처리 전에 매입조회 파일이 먼저 정상 로드되어야 합니다.")
    if "회계월" not in base_df.columns or base_df["회계월"].dropna().empty:
        raise KeyError("매입조회 파일에 '회계월' 컬럼이 없습니다.")

    current_month = int(base_df["회계월"].dropna().iloc[0])
    previous_month = 12 if current_month == 1 else current_month - 1
    col_name = f"{previous_month}월 기말여부"

    if col_name not in df.columns:
        raise KeyError(f"기초재고 파일에 '{col_name}' 컬럼이 없습니다.")

    filtered_df = df[df[col_name] == 1].copy()
    return df, filtered_df


def filter_purchase_inquiry(df_purchase, df_inventory_all):
    """매입조회에서 기초재고/전월마스터 선매입 상품ID 제외."""
    if df_purchase is None or df_purchase.empty:
        return df_purchase

    df_purchase = df_purchase.copy()

    if df_inventory_all is None or df_inventory_all.empty:
        return df_purchase
    if "상품ID" not in df_purchase.columns:
        return df_purchase
    if "상품ID" not in df_inventory_all.columns:
        return df_purchase

    prepaid_columns = [
        column
        for column in ("선매입 여부", "선매입여부", "선매입")
        if column in df_inventory_all.columns
    ]
    if not prepaid_columns:
        return df_purchase

    prepaid_mask = pd.Series(False, index=df_inventory_all.index)
    for column in prepaid_columns:
        values = df_inventory_all[column].astype(str).str.strip()
        prepaid_mask = prepaid_mask | values.eq("선매입") | _is_flag_one(df_inventory_all[column])

    excluded_ids = set(
        df_inventory_all.loc[prepaid_mask, "상품ID"]
        .dropna().astype(str).str.strip()
    )
    excluded_ids.discard("")
    if not excluded_ids:
        return df_purchase

    purchase_ids = df_purchase["상품ID"].fillna("").astype(str).str.strip()
    return df_purchase[~purchase_ids.isin(excluded_ids)].copy()


# ============================================================
# 4. 차량 상세 / lookup 빌더
# ============================================================

def _build_detail_frame(df, date_column):
    """매입조회/기초재고/전체상품조회/위탁수불부 → 상세 컬럼 통일."""
    columns = ["상품ID", *DETAIL_COLUMNS]
    if df is None or df.empty or "상품ID" not in df.columns:
        return pd.DataFrame(columns=columns)

    temp = _strip_columns(df)
    temp["상품ID"] = temp["상품ID"].astype(str).str.strip()
    temp = temp[temp["상품ID"] != ""].copy()
    if temp.empty:
        return pd.DataFrame(columns=columns)

    detail = pd.DataFrame({"상품ID": temp["상품ID"]})
    detail["신번호"] = _get_first_existing_column(temp, ["차량번호", "신번호"])
    detail["구번호"] = _get_first_existing_column(temp, ["이전차량번호", "구번호", "이전차량번호1"])
    detail["차대번호"] = _get_first_existing_column(temp, ["차대번호"])
    detail["차종"] = _get_first_existing_column(temp, ["차종"])
    detail["차명"] = _get_first_existing_column(temp, ["차명", "차량명"])
    detail["반납일자"] = _get_first_existing_column(temp, ["반납일자"])
    date_candidates = [date_column, "매입일자"] if date_column is not None else ["매입일자"]
    detail["매입일자"] = _get_first_existing_column(temp, date_candidates)
    detail["분류1"] = _get_first_existing_column(temp, ["신매입유형1", "매입유형-분류1", "분류1"])
    detail["분류2"] = _get_first_existing_column(temp, ["신매입유형2", "매입유형-분류2", "분류2"])
    detail["분류3"] = _get_first_existing_column(temp, ["신매입유형3", "매입유형-분류3", "분류3"])
    detail["분류4"] = _get_first_existing_column(temp, ["채널", "매입채널", "매입유형-분류4", "분류4"])

    return detail.drop_duplicates(subset=["상품ID"], keep="first").reset_index(drop=True)


def _build_sales_vehicle_detail_frame(df):
    """검사/정비매출용 - 신번호/구번호만 채우는 단순 상세."""
    columns = ["상품ID", *DETAIL_COLUMNS]
    if df is None or df.empty or "상품ID" not in df.columns:
        return pd.DataFrame(columns=columns)

    temp = _strip_columns(df)
    temp["상품ID"] = temp["상품ID"].astype(str).str.strip()
    temp = temp[temp["상품ID"] != ""].copy()
    if temp.empty:
        return pd.DataFrame(columns=columns)

    vehicle_numbers = _get_first_existing_column(temp, ["차량번호", "신번호"])
    detail = pd.DataFrame({"상품ID": temp["상품ID"]})
    detail["신번호"] = vehicle_numbers
    detail["구번호"] = vehicle_numbers
    for column in DETAIL_COLUMNS:
        if column not in ("신번호", "구번호"):
            detail[column] = ""

    return detail.drop_duplicates(subset=["상품ID"], keep="first").reset_index(drop=True)


def _append_vehicle_details(merged, dfs):
    """출처별 차량 상세를 머지하고 우선순위에 따라 단일 컬럼으로 결정."""
    sources = {
        "매입": _build_detail_frame(dfs.get("매입조회"), "계산서일자"),
        "기초": _build_detail_frame(dfs.get("기초재고"), "계산서일자"),
        "전체": _build_detail_frame(dfs.get("전체상품조회"), "매입세금계산서일자"),
        "위탁수불": _build_detail_frame(dfs.get("위탁수불부_전체"), None),
        "검사": _build_sales_vehicle_detail_frame(dfs.get("검사매출")),
        "정비": _build_sales_vehicle_detail_frame(dfs.get("정비매출")),
    }

    helper_columns = []
    for prefix, detail_df in sources.items():
        renamed = detail_df.rename(
            columns={column: f"{prefix}_{column}" for column in DETAIL_COLUMNS}
        )
        merged = merged.merge(renamed, on="상품ID", how="left")
        helper_columns.extend(f"{prefix}_{column}" for column in DETAIL_COLUMNS)

    use_inventory = merged["_출처"].eq("기초재고")
    use_inspection = merged["_출처"].eq("검사매출")
    use_maintenance = merged["_출처"].eq("정비매출")
    use_product_master = merged["_출처"].eq("전체상품조회")
    use_consignment = merged["_출처"].isin(["위탁수불부_기초", "위탁수불부_입고"])

    for column in DETAIL_COLUMNS:
        # object dtype 으로 강제 캐스팅 — combine_first 가 dtype 자동변환 시도하다
        # to_datetime 으로 터지는 문제(pandas 신버전) 회피
        purchase_v = merged[f"매입_{column}"].replace("", pd.NA).astype("object")
        inventory_v = merged[f"기초_{column}"].replace("", pd.NA).astype("object")
        product_master_v = merged[f"전체_{column}"].replace("", pd.NA).astype("object")
        consignment_v = merged[f"위탁수불_{column}"].replace("", pd.NA).astype("object")
        inspection_v = merged[f"검사_{column}"].fillna("").astype("object")
        maintenance_v = merged[f"정비_{column}"].fillna("").astype("object")

        default_v = (
            purchase_v.combine_first(inventory_v)
                      .combine_first(product_master_v)
                      .combine_first(consignment_v)
        )
        inventory_result = (
            inventory_v.combine_first(consignment_v)
                       .combine_first(product_master_v)
                       .combine_first(purchase_v)
        )
        product_master_result = (
            product_master_v.combine_first(purchase_v)
                            .combine_first(consignment_v)
                            .combine_first(inventory_v)
        )
        consignment_result = (
            consignment_v.combine_first(product_master_v)
                         .combine_first(purchase_v)
                         .combine_first(inventory_v)
        )

        result = default_v
        result = result.where(~use_product_master, product_master_result)
        result = result.where(~use_consignment, consignment_result)
        result = result.where(~use_inventory, inventory_result)
        result = result.where(~use_inspection, inspection_v)
        result = result.where(~use_maintenance, maintenance_v)
        merged[column] = result.fillna("")

    return merged.drop(columns=helper_columns)


def _build_detail_lookup(detail_df):
    """차량번호(신/구) → 상품ID, 상품ID → 분류1 lookup."""
    empty_lookup = {
        "primary_vehicle_rows": [],
        "secondary_vehicle_rows": [],
        "product_type_by_id": {},
    }
    if detail_df is None or detail_df.empty:
        return empty_lookup

    detail = _strip_columns(detail_df)
    if any(column not in detail.columns for column in ("신번호", "상품ID")):
        return empty_lookup

    detail = detail.copy()
    detail["_신번호_lookup"] = detail["신번호"].apply(_normalize_lookup_value)
    detail["_구번호_lookup"] = (
        detail["구번호"].apply(_normalize_lookup_value)
        if "구번호" in detail.columns else ""
    )
    detail["_상품ID_lookup"] = detail["상품ID"].apply(_normalize_lookup_value)

    primary = (
        detail.loc[detail["_신번호_lookup"] != "", ["_신번호_lookup", "상품ID"]]
              .dropna(subset=["상품ID"]).values.tolist()
    )
    secondary = (
        detail.loc[detail["_구번호_lookup"] != "", ["_구번호_lookup", "상품ID"]]
              .dropna(subset=["상품ID"]).values.tolist()
    )

    product_type_by_id = {}
    if "분류1" in detail.columns:
        type_lookup = (
            detail.loc[detail["_상품ID_lookup"] != "", ["_상품ID_lookup", "분류1"]]
                  .drop_duplicates(subset=["_상품ID_lookup"], keep="first")
        )
        product_type_by_id = dict(zip(type_lookup["_상품ID_lookup"], type_lookup["분류1"]))

    return {
        "primary_vehicle_rows": primary,
        "secondary_vehicle_rows": secondary,
        "product_type_by_id": product_type_by_id,
    }


def _lookup_product_id_in_lookup(lookup_car_number, detail_lookup):
    for detail_car_number, product_id in reversed(detail_lookup["primary_vehicle_rows"]):
        if lookup_car_number in detail_car_number:
            return product_id
    for detail_car_number, product_id in reversed(detail_lookup["secondary_vehicle_rows"]):
        if lookup_car_number in detail_car_number:
            return product_id
    return ""


def _lookup_product_id_by_car_number(car_number, detail_lookup, fallback_lookup=None):
    lookup_car_number = _normalize_lookup_value(car_number)
    if lookup_car_number == "":
        return ""

    product_id = _lookup_product_id_in_lookup(lookup_car_number, detail_lookup)
    if product_id != "" or fallback_lookup is None:
        return product_id
    return _lookup_product_id_in_lookup(lookup_car_number, fallback_lookup)


def _build_product_ledger_lookup(product_ledger_df):
    """상품원장 차량번호 → 상품ID."""
    if (
        product_ledger_df is None
        or product_ledger_df.empty
        or "차량번호" not in product_ledger_df.columns
        or "상품ID" not in product_ledger_df.columns
    ):
        return {}

    ledger = _strip_columns(product_ledger_df).copy()
    ledger["_차량번호_lookup"] = ledger["차량번호"].apply(_normalize_lookup_value)
    ledger = ledger[ledger["_차량번호_lookup"] != ""].copy()
    if ledger.empty:
        return {}

    ledger = ledger.drop_duplicates(subset=["_차량번호_lookup"], keep="last")
    return dict(zip(ledger["_차량번호_lookup"], ledger["상품ID"]))


def _build_product_id_lookup_by_column(detail_df, lookup_column):
    """detail_df 의 lookup_column → 상품ID."""
    if (
        detail_df is None
        or detail_df.empty
        or lookup_column not in detail_df.columns
        or "상품ID" not in detail_df.columns
    ):
        return {}

    detail = _strip_columns(detail_df).copy()
    detail["_lookup_key"] = detail[lookup_column].apply(_normalize_lookup_value)
    detail = detail[detail["_lookup_key"] != ""].copy()
    if detail.empty:
        return {}

    detail = detail.drop_duplicates(subset=["_lookup_key"], keep="last")
    return dict(zip(detail["_lookup_key"], detail["상품ID"]))


# ============================================================
# 5. 상품ID 모음
# ============================================================

def collect_product_ids(dfs, settlement_year=None, settlement_month=None):
    base_ids = []
    for key in PRODUCT_ID_SOURCE_KEYS:
        df = dfs.get(key)
        if df is None or df.empty or "상품ID" not in df.columns:
            continue

        temp = df[["상품ID"]].copy()
        temp["_출처"] = key
        temp["_입고구분"] = ""
        if key == "매입조회" and "입고구분" in df.columns:
            temp["_입고구분"] = df["입고구분"].astype(str).str.strip()

        temp["상품ID"] = temp["상품ID"].astype(str).str.strip()
        temp = temp[temp["상품ID"] != ""].copy()
        base_ids.append(temp)

    if not base_ids:
        return pd.DataFrame(columns=PRODUCT_ID_COLUMNS)

    merged = pd.concat(base_ids, ignore_index=True).reset_index(drop=True)

    # 입고/기초 플래그
    for column in ("기초재고", "정상입고", "타처입고"):
        merged[column] = 0
    merged["선매입여부"] = ""

    purchase_rows = merged["_출처"].eq("매입조회")
    merged.loc[merged["_출처"].eq("기초재고"), "기초재고"] = 1
    merged.loc[merged["_출처"].eq("위탁수불부_기초"), "기초재고"] = 1
    merged.loc[purchase_rows & merged["_입고구분"].eq("정상입고"), "정상입고"] = 1
    merged.loc[purchase_rows & merged["_입고구분"].eq("타처입고"), "타처입고"] = 1
    merged.loc[merged["_출처"].eq("전체상품조회"), "정상입고"] = 1
    merged.loc[merged["_출처"].eq("전체상품조회"), "선매입여부"] = "선매입"
    merged.loc[merged["_출처"].eq("위탁수불부_입고"), "정상입고"] = 1

    # 매출구분 / 당사/타사
    merged["구분2"] = ""
    internal_sales = merged[["기초재고", "정상입고", "타처입고"]].eq(1).any(axis=1)
    merged.loc[internal_sales, "구분2"] = "사내매출"
    merged.loc[
        merged["_출처"].isin(["위탁수불부_기초", "위탁수불부_입고"]), "구분2"
    ] = "위탁매출"
    merged.loc[merged["_출처"].eq("검사매출"), "구분2"] = "검사매출"
    merged.loc[merged["_출처"].eq("정비매출"), "구분2"] = "정비매출"
    merged["구분1"] = merged["구분2"].eq("사내매출").map({True: "당사차량", False: "타사차량"})

    # 차량 상세 + 매입연/월
    merged = _append_vehicle_details(merged, dfs)
    parsed_purchase_date = pd.to_datetime(merged["매입일자"], format="mixed", errors="coerce")
    internal_sales = merged["구분2"].eq("사내매출")
    merged["매입연도"] = parsed_purchase_date.dt.year.astype("Int64").where(internal_sales, pd.NA)
    merged["매입월"] = parsed_purchase_date.dt.month.astype("Int64").where(internal_sales, pd.NA)
    merged["회계연도"] = int(settlement_year) if settlement_year is not None else pd.NA
    merged["회계월"] = int(settlement_month) if settlement_month is not None else pd.NA

    merged = merged.drop(columns=["_출처", "_입고구분"])
    merged = merged.rename(columns={"구분1": "당사/타사", "구분2": "매출구분"})
    merged = merged[PRODUCT_ID_COLUMNS]
    return _sort_product_id_df(merged)


# ============================================================
# 6. 전처리 - 매입원가
# ============================================================

# ----- 상품원장 -----

def extract_reference(text):
    text = str(text) if pd.notna(text) else ""
    patterns = [
        r"C\d{11}_\d{2,3}[^\d]\d{4}",
        r"C\d{11}_[^\d]{3}",
        r"C\d{11}_[^\d]{2}\d{2,3}[^\d]\d{4}",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group()
    return ""


# 원가구분 키워드 → 라벨 (순서 중요: 위에서부터 매칭)
_COST_KEYWORD_RULES = [
    (("매출원가", "재공품", "상품평가충당금"), "결산"),
    (("오류",), "매입수수료"),
    (("초과운행",), "초과운행"),
    (("계약만기 도래분(반납)",), "페이백(반납)"),
    (("계약만기 도래분(미반납)",), "페이백(미반납)"),
    (("폐자원",), "폐자원공제"),
    (("취득세", "취등록세"), "취득세"),
    (("선매입",), "상품매입액"),
    (
        (
            "피알앤디컴퍼니", "경매장", "인품", "엔카", "중개", "알선", "매입",
            "소개수수료", "헤이딜러", "매입수수료", "매입 수수료",
            "낙찰수수료", "낙찰 수수료",
        ),
        "매입수수료",
    ),
    (("(상품->건설중인자산)", "상품->자산"), "자산출고"),
    (("상품전환",), "타처입고"),
]


def classify_cost(row):
    text = str(row["적요"]) if pd.notna(row["적요"]) else ""
    for keywords, label in _COST_KEYWORD_RULES:
        if any(keyword in text for keyword in keywords):
            return label

    reference = row["참고"]
    if pd.notna(reference) and reference not in ("", 0):
        return "상품매입액"
    return ""


def extract_car_number(row):
    text = str(row["적요"]) if pd.notna(row["적요"]) else ""
    cost_type = row["원가구분"]

    if cost_type == "결산":
        return "결산"
    if text.endswith("지게차"):
        return "지게차"
    if cost_type in ("페이백(반납)", "페이백(미반납)", "폐자원공제", "초과운행"):
        return "확인필요"

    region_prefix = (
        "서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주"
    )
    match = re.search(rf"(?:{region_prefix})?\d{{2,3}}[가-힣]\d{{4}}", text)
    if not match:
        match = re.search(r"\d{2,3}[^\d]\d{4}", text)
    return match.group() if match else ""


def _lookup_product_id_from_product_id_df(car_number, accounting_month, product_id_df):
    """product_id_df 에서 상품원장 행에 매칭되는 상품ID 반환.

    조건 (엑셀 수식 동등):
      매출구분 == "사내매출"
      AND (신번호 == 차량번호 OR 구번호 == 차량번호)
      AND 회계월 == accounting_month
    매칭 행 없으면 빈 문자열 반환.
    """
    if product_id_df is None or product_id_df.empty:
        return ""
    if not car_number:
        return ""

    if "매출구분" not in product_id_df.columns:
        return ""

    sales_type = product_id_df["매출구분"].astype(str).str.strip()
    mask = sales_type.eq("사내매출")

    if "회계월" in product_id_df.columns and accounting_month is not None:
        try:
            month_int = int(accounting_month)
        except (TypeError, ValueError):
            month_int = None
        if month_int is not None:
            month_series = pd.to_numeric(product_id_df["회계월"], errors="coerce")
            mask = mask & month_series.eq(month_int)

    car_norm = _normalize_lookup_value(car_number)
    if "신번호" in product_id_df.columns:
        new_no = product_id_df["신번호"].apply(_normalize_lookup_value)
    else:
        new_no = pd.Series([""] * len(product_id_df), index=product_id_df.index)
    if "구번호" in product_id_df.columns:
        old_no = product_id_df["구번호"].apply(_normalize_lookup_value)
    else:
        old_no = pd.Series([""] * len(product_id_df), index=product_id_df.index)

    mask = mask & (new_no.eq(car_norm) | old_no.eq(car_norm))

    matched = product_id_df[mask]
    if matched.empty:
        return ""
    if "상품ID" not in matched.columns:
        return ""
    pid = matched["상품ID"].iloc[0]
    return "" if pd.isna(pid) else str(pid)


def _append_product_ledger_purchase_columns(df, detail_df, inventory_all_df=None, product_id_df=None):
    """상품원장에 상품ID, 매입유형 컬럼 부여.

    상품ID 매칭 우선순위:
      1) 차량번호 == "지게차" → 적요 앞 12글자
      2) 그 외 → product_id_df 에서 (사내매출 + 신/구번호 + 회계월) 매칭
    """
    detail_lookup_df = detail_df
    if detail_lookup_df is not None and not detail_lookup_df.empty:
        detail_lookup_df = _strip_columns(detail_lookup_df)
        if "당사/타사" in detail_lookup_df.columns:
            detail_lookup_df = detail_lookup_df[
                detail_lookup_df["당사/타사"].astype(str).str.strip().eq("당사차량")
            ].copy()

    detail_lookup = _build_detail_lookup(detail_lookup_df)
    inventory_detail_df = _build_detail_frame(inventory_all_df, "위탁등록일자")
    inventory_lookup = _build_detail_lookup(inventory_detail_df)

    # 차량번호+회계월 단위 캐시 (반복 매칭 비용 절감)
    lookup_cache = {}
    product_ids = []
    for _, row in df.iterrows():
        car_number = str(row["차량번호"]).strip() if pd.notna(row["차량번호"]) else ""

        if car_number == "지게차":
            summary_text = str(row["적요"]) if pd.notna(row["적요"]) else ""
            product_id = summary_text[:12]
        else:
            accounting_month = row.get("회계월", None)
            cache_key = (car_number, accounting_month)
            if cache_key not in lookup_cache:
                lookup_cache[cache_key] = _lookup_product_id_from_product_id_df(
                    car_number, accounting_month, product_id_df,
                )
            product_id = lookup_cache[cache_key]
        product_ids.append(product_id)
    df["상품ID"] = product_ids

    product_type_by_id = inventory_lookup["product_type_by_id"].copy()
    product_type_by_id.update(detail_lookup["product_type_by_id"])
    df["매입유형"] = [
        product_type_by_id.get(_normalize_lookup_value(pid), "")
        for pid in df["상품ID"]
    ]
    return df


def preprocess_product_ledger(
    file,
    detail_df=None,
    inventory_all_df=None,
    settlement_year=None,
    settlement_month=None,
    product_id_df=None,
):
    """상품원장"""
    file.seek(0)
    df = _strip_columns(pd.read_excel(file))
    df = df[~df["회계일자"].isin(["월계", "누계", "전일이월"])].copy()

    if "작성사원명" in df.columns:
        df = df[df["작성사원명"].astype(str).str.strip() != "김겸윤"].copy()

    df["회계일자"] = pd.to_datetime(df["회계일자"], format="mixed", errors="coerce")
    df = df[df["회계일자"].notna()].copy()
    df["회계연도"] = df["회계일자"].dt.year
    df["회계월"] = df["회계일자"].dt.month
    df["회계일자"] = df["회계일자"].dt.date

    if settlement_year is not None and settlement_month is not None:
        df = df[
            (df["회계연도"] == int(settlement_year))
            & (df["회계월"] == int(settlement_month))
        ].copy()

    df["참고"] = df["적요"].apply(extract_reference)
    if "원가구분" not in df.columns:
        df["원가구분"] = df.apply(classify_cost, axis=1)
    df["차량번호"] = df.apply(extract_car_number, axis=1)

    df["차변"] = pd.to_numeric(df["차변"], errors="coerce").fillna(0)
    df["대변"] = pd.to_numeric(df["대변"], errors="coerce").fillna(0)
    df["abs_v"] = df["차변"].abs()
    df["seq"] = df.groupby(
        ["회계연도", "회계월", "차량번호", "abs_v", df["차변"] > 0]
    ).cumcount()

    canceled = (
        df.groupby(["회계연도", "회계월", "차량번호", "abs_v", "seq"])["차변"]
          .transform("count") > 1
    )
    df["상태"] = np.where(canceled, "취소", "")
    df.loc[df["상태"].eq("") & df["원가구분"].eq("결산"), "상태"] = "결산"
    df.drop(columns=["abs_v", "seq"], inplace=True)
    df["금액"] = df["차변"] - df["대변"]

    if "잔액" in df.columns and "작성일자" in df.columns:
        columns_to_remove = df.loc[:, "잔액":"작성일자"].columns
        df.drop(columns=columns_to_remove, inplace=True)

    return _append_product_ledger_purchase_columns(
        df, detail_df, inventory_all_df, product_id_df=product_id_df,
    )


# ----- 폐자원 -----

def _lookup_row_priority(row, value_column):
    value = "" if pd.isna(row[value_column]) else str(row[value_column]).strip()
    if value == "":
        return 0

    prepaid_value = ""
    for column in ("선매입여부", "선매입"):
        if column in row.index:
            prepaid_value = str(row[column]).strip()
            if prepaid_value == "선매입":
                break

    if value == "선매입" or prepaid_value == "선매입":
        return 2
    return 1


def _store_chassis_month_lookup_value(lookup, priorities, key, value, priority):
    if key not in lookup or priority > priorities.get(key, -1):
        lookup[key] = value
        priorities[key] = priority


def _merge_chassis_month_note_lookups(*lookups):
    """비고 lookup 병합. 빈 값보다 실제 값, 특히 '선매입'을 우선한다."""
    merged = {}
    for lookup in lookups:
        if not lookup:
            continue
        for key, raw_value in lookup.items():
            value = "" if pd.isna(raw_value) else str(raw_value).strip()
            if value == "":
                continue
            current = merged.get(key, "")
            if current == "" or value == "선매입":
                merged[key] = value
    return merged


def _build_chassis_month_value_lookup(
    df, value_column, year_column="회계연도", month_column="회계월",
):
    """detail/master df 에서 (차대번호, 연도, 월) → value_column lookup 생성."""
    if df is None or df.empty:
        return {}
    if "차대번호" not in df.columns or month_column not in df.columns:
        return {}
    if value_column not in df.columns:
        return {}

    work = _strip_columns(df)
    work["_차대번호_lookup"] = work["차대번호"].apply(_normalize_lookup_value)
    work["_연도_lookup"] = (
        pd.to_numeric(work[year_column], errors="coerce")
        if year_column in work.columns else pd.NA
    )
    work["_월_lookup"] = pd.to_numeric(work[month_column], errors="coerce")
    work = work[
        work["_차대번호_lookup"].ne("")
        & work["_월_lookup"].notna()
    ].copy()
    if work.empty:
        return {}

    lookup = {}
    priorities = {}
    for _, row in work.iterrows():
        year_value = row["_연도_lookup"]
        year_key = int(year_value) if pd.notna(year_value) else None
        month_key = int(row["_월_lookup"])
        key = (row["_차대번호_lookup"], year_key, month_key)
        value = "" if pd.isna(row[value_column]) else str(row[value_column]).strip()
        priority = _lookup_row_priority(row, value_column)
        _store_chassis_month_lookup_value(lookup, priorities, key, value, priority)
    return lookup


@functools.lru_cache(maxsize=2)
def _read_cost_summary_master_cached(path, mtime, size):
    """(path, mtime, size)가 같으면 재사용 — 같은 실행/리런 내 중복 읽기 방지.

    maxsize 를 작게 유지해 큰 마스터 DataFrame이 여러 버전 메모리에 쌓이지 않게 함."""
    try:
        return _strip_columns(pd.read_excel(path, sheet_name="최종원가마스터"))
    except Exception:
        try:
            return _strip_columns(pd.read_excel(path))
        except Exception:
            return None


def _load_latest_cost_summary_master():
    """앱 폴더의 최신 cost_summary_*.xlsx 를 읽어 반환. 없거나 실패하면 빈 df."""
    import glob
    import os

    here = os.path.dirname(os.path.abspath(__file__))
    paths = [
        path for path in glob.glob(os.path.join(here, "cost_summary_*.xlsx"))
        if os.path.exists(path) and os.path.getsize(path) > 0
    ]
    if not paths:
        return pd.DataFrame()

    paths.sort(key=lambda path: os.path.basename(path), reverse=True)
    for path in paths:
        stat = os.stat(path)
        result = _read_cost_summary_master_cached(path, stat.st_mtime, stat.st_size)
        if result is not None:
            return result
    return pd.DataFrame()


def _find_prepaid_column(df):
    """선매입 여부 컬럼명 후보 반환."""
    if df is None or df.empty:
        return None
    for column in ("선매입여부", "선매입"):
        if column in df.columns:
            return column
    return None


def _find_purchase_year_column(df):
    if df is None or df.empty:
        return "매입연도"
    for column in ("매입연도", "매입년도"):
        if column in df.columns:
            return column
    return "매입연도"


def _lookup_chassis_month_value(lookup, chassis_value, year_value, month_value):
    chassis_key = _normalize_lookup_value(chassis_value)
    if not chassis_key:
        return ""
    if pd.isna(year_value):
        year_key = None
    else:
        try:
            year_key = int(year_value)
        except (TypeError, ValueError):
            year_key = None
    if pd.isna(month_value):
        return ""
    try:
        month_key = int(month_value)
    except (TypeError, ValueError):
        return ""
    value = lookup.get((chassis_key, year_key, month_key), "")
    if value != "":
        return value
    return lookup.get((chassis_key, None, month_key), "")


def preprocess_waste_resource_file(file, product_ledger_df=None, detail_df=None):
    sheets = _load_excel_sheets(file)
    product_id_by_chassis_month = _build_chassis_month_value_lookup(detail_df, "상품ID")
    product_id_by_chassis = _build_product_id_lookup_by_column(detail_df, "차대번호")

    master_df = _load_latest_cost_summary_master()
    prepaid_column = _find_prepaid_column(master_df)
    detail_prepaid_column = _find_prepaid_column(detail_df)
    master_purchase_year_column = _find_purchase_year_column(master_df)
    detail_purchase_year_column = _find_purchase_year_column(detail_df)
    prepaid_by_chassis_month = _merge_chassis_month_note_lookups(
        _build_chassis_month_value_lookup(
            master_df, prepaid_column,
            year_column=master_purchase_year_column,
            month_column="매입월",
        )
        if prepaid_column is not None else {},
        _build_chassis_month_value_lookup(
            detail_df, detail_prepaid_column,
            year_column=detail_purchase_year_column,
            month_column="매입월",
        )
        if detail_prepaid_column is not None else {},
    )

    processed_sheets = {}
    for sheet_name, df in sheets.items():
        if "구분" in df.columns:
            df = df[df["구분"].astype(str).str.strip().isin(["영수증", "계산서"])].copy()

        # 제외대상 (세액공제액 < 0)
        tax_credit_column = next(
            (c for c in WASTE_RESOURCE_AMOUNT_COLUMNS if c in df.columns), None
        )
        if tax_credit_column is not None:
            tax_credit = pd.to_numeric(df[tax_credit_column], errors="coerce").fillna(0)
            df["제외대상"] = np.where(tax_credit < 0, 1, 0)
        else:
            df["제외대상"] = 0

        # 매입일자 기준 매입년도/매입월 생성
        if "매입일자" in df.columns:
            parsed_purchase_date = pd.to_datetime(
                df["매입일자"], format="mixed", errors="coerce"
            )
            df["매입년도"] = parsed_purchase_date.dt.year.astype("Int64")
            df["매입월"] = parsed_purchase_date.dt.month.astype("Int64")
        else:
            df["매입년도"] = pd.NA
            df["매입월"] = pd.NA

        purchase_month = pd.to_numeric(df["매입월"], errors="coerce")
        purchase_year = pd.to_numeric(df["매입년도"], errors="coerce")
        previous_month = (purchase_month - 1).where(purchase_month.ne(1), 12)
        previous_year = purchase_year.where(purchase_month.ne(1), purchase_year - 1)

        # 비고: 기존 마스터/원가대상의 같은 차대번호 + 매입월 선매입여부 조회
        df["비고"] = ""
        if "차대번호" in df.columns and prepaid_by_chassis_month:
            note_values = []
            for chassis_value, year_value, month_value in zip(
                df["차대번호"], purchase_year, purchase_month
            ):
                if pd.isna(month_value):
                    note_values.append("")
                else:
                    note_values.append(
                        _lookup_chassis_month_value(
                            prepaid_by_chassis_month,
                            chassis_value,
                            year_value,
                            month_value,
                        )
                    )
            df["비고"] = note_values

        # 회계연도/회계월: 선매입이면 매입월의 직전월, 아니면 매입월 그대로
        df["회계연도"] = pd.to_numeric(df["매입년도"], errors="coerce").astype("Int64")
        df["회계월"] = purchase_month.astype("Int64")
        prepaid_rows = df["비고"].astype(str).str.strip().eq("선매입")
        df.loc[prepaid_rows, "회계연도"] = previous_year[prepaid_rows].astype("Int64")
        df.loc[prepaid_rows, "회계월"] = previous_month[prepaid_rows].astype("Int64")

        # 상품ID: 차대번호가 없거나 제외대상이면 빈 값, 아니면 원가대상의 같은 회계월 차대번호에서 조회
        df["상품ID"] = ""
        if "차대번호" in df.columns:
            excluded_rows = _is_flag_one(df["제외대상"])
            product_ids = []
            for chassis_value, accounting_year, accounting_month, excluded in zip(
                df["차대번호"], df["회계연도"], df["회계월"], excluded_rows
            ):
                if excluded:
                    product_ids.append("")
                    continue
                product_id = _lookup_chassis_month_value(
                    product_id_by_chassis_month,
                    chassis_value,
                    accounting_year,
                    accounting_month,
                )
                if not product_id and not product_id_by_chassis_month:
                    product_id = product_id_by_chassis.get(
                        _normalize_lookup_value(chassis_value), ""
                    )
                product_ids.append(product_id)
            df["상품ID"] = product_ids

        # 컬럼 순서: 첫 컬럼 다음에 [제외대상, 매입년도, 매입월, 비고, 회계연도, 회계월] 삽입
        ordered = [
            c for c in ("제외대상", "매입년도", "매입월", "비고", "회계연도", "회계월")
            if c in df.columns
        ]
        rest = [c for c in df.columns if c not in ordered]
        df = df[rest[:1] + ordered + rest[1:]]

        processed_sheets[sheet_name] = df

    return processed_sheets


# ----- 페이백 -----

def _parse_year_month(value):
    if pd.isna(value):
        return pd.NA, pd.NA
    text = str(value).strip()
    match = re.search(r"(\d{4})\D?(\d{1,2})", text)
    if not match:
        return pd.NA, pd.NA
    return int(match.group(1)), int(match.group(2))


def preprocess_payback_file(file, detail_df=None, settlement_year=None, settlement_month=None):
    sheets = _load_excel_sheets(file)
    detail_lookup = _build_detail_lookup(detail_df)

    processed_sheets = {}
    for sheet_name, df in sheets.items():
        # 결산연/월 필터
        if "연도월" in df.columns:
            parsed_period = df["연도월"].apply(_parse_year_month)
            df["회계연도"] = parsed_period.apply(lambda v: v[0]).astype("Int64")
            df["회계월"] = parsed_period.apply(lambda v: v[1]).astype("Int64")
            if settlement_year is not None and settlement_month is not None:
                df = df[
                    (df["회계연도"] == int(settlement_year))
                    & (df["회계월"] == int(settlement_month))
                ].copy()
        elif settlement_year is not None and settlement_month is not None:
            df = df.iloc[0:0].copy()

        original_product_ids = (
            df["상품ID"].copy()
            if "상품ID" in df.columns
            else pd.Series([""] * len(df), index=df.index)
        )

        if "차량번호" in df.columns:
            product_ids = []
            for index, row in df.iterrows():
                car_number = str(row["차량번호"]).strip() if pd.notna(row["차량번호"]) else ""
                normalized_car_number = _normalize_lookup_value(car_number)

                if len(normalized_car_number) == 12:
                    product_id = normalized_car_number
                elif car_number == "지게차":
                    product_id = original_product_ids.loc[index]
                else:
                    product_id = _lookup_product_id_by_car_number(car_number, detail_lookup)
                product_ids.append(product_id)
            df["상품ID"] = product_ids
        elif "상품ID" not in df.columns:
            df["상품ID"] = ""

        processed_sheets[sheet_name] = df

    return processed_sheets


# ----- 일반 원가 파일 -----

def preprocess_cost_file(file):
    """기타 원가 시트 (특별 처리 없음)."""
    return _load_excel_sheets(file)


# ============================================================
# 7. 전처리 - 제조원가
# ============================================================

def _build_vehicle_sales_count_lookup(detail_df):
    """매출구분 × {신번호, 구번호} → 차량별 카운트."""
    lookup = {
        sales_type: {"신번호": {}, "구번호": {}}
        for sales_type in MATERIAL_SALES_COLUMNS
    }
    if detail_df is None or detail_df.empty:
        return lookup

    detail = _strip_columns(detail_df)
    required_columns = ["매출구분", "신번호", "구번호"]
    if any(column not in detail.columns for column in required_columns):
        return lookup

    detail = detail.copy()
    detail["_신번호_lookup"] = detail["신번호"].apply(_normalize_lookup_value)
    detail["_구번호_lookup"] = detail["구번호"].apply(_normalize_lookup_value)
    detail["매출구분"] = detail["매출구분"].astype(str).str.strip()

    for sales_type in MATERIAL_SALES_COLUMNS:
        sales_rows = detail[detail["매출구분"].eq(sales_type)]
        new_counts = sales_rows.loc[
            sales_rows["_신번호_lookup"] != "", "_신번호_lookup"
        ].value_counts()
        old_counts = sales_rows.loc[
            sales_rows["_구번호_lookup"] != "", "_구번호_lookup"
        ].value_counts()
        lookup[sales_type]["신번호"] = new_counts.to_dict()
        lookup[sales_type]["구번호"] = old_counts.to_dict()

    return lookup


def _get_row_sales_count(row, sales_type, sales_count_lookup):
    vehicle_key = _normalize_lookup_value(row["차량번호"] if "차량번호" in row.index else "")
    old_vehicle_key = _normalize_lookup_value(row["구차량번호"] if "구차량번호" in row.index else "")
    return (
        sales_count_lookup[sales_type]["신번호"].get(vehicle_key, 0)
        + sales_count_lookup[sales_type]["구번호"].get(old_vehicle_key, 0)
    )


def _build_direct_expense_sales_count_lookup(detail_df):
    """직접경비 전용: 매출구분 × {신번호, 구번호, 상품ID} → 카운트.

    직접경비는 차량번호 컬럼 하나만 있으므로, 차량번호를 신번호/구번호 양쪽과
    매칭하고 상품아이디를 상품ID 와도 매칭하기 위한 lookup.
    """
    lookup = {
        sales_type: {"신번호": {}, "구번호": {}, "상품ID": {}}
        for sales_type in MATERIAL_SALES_COLUMNS
    }
    if detail_df is None or detail_df.empty:
        return lookup

    detail = _strip_columns(detail_df)
    if "매출구분" not in detail.columns:
        return lookup

    detail = detail.copy()
    detail["매출구분"] = detail["매출구분"].astype(str).str.strip()
    if "신번호" in detail.columns:
        detail["_신번호_lookup"] = detail["신번호"].apply(_normalize_lookup_value)
    else:
        detail["_신번호_lookup"] = ""
    if "구번호" in detail.columns:
        detail["_구번호_lookup"] = detail["구번호"].apply(_normalize_lookup_value)
    else:
        detail["_구번호_lookup"] = ""
    if "상품ID" in detail.columns:
        detail["_상품ID_lookup"] = detail["상품ID"].apply(_normalize_lookup_value)
    else:
        detail["_상품ID_lookup"] = ""

    for sales_type in MATERIAL_SALES_COLUMNS:
        sales_rows = detail[detail["매출구분"].eq(sales_type)]
        lookup[sales_type]["신번호"] = (
            sales_rows.loc[sales_rows["_신번호_lookup"] != "", "_신번호_lookup"]
            .value_counts().to_dict()
        )
        lookup[sales_type]["구번호"] = (
            sales_rows.loc[sales_rows["_구번호_lookup"] != "", "_구번호_lookup"]
            .value_counts().to_dict()
        )
        lookup[sales_type]["상품ID"] = (
            sales_rows.loc[sales_rows["_상품ID_lookup"] != "", "_상품ID_lookup"]
            .value_counts().to_dict()
        )

    return lookup


def _get_direct_expense_sales_count(
    sales_type, lookup, car_number, source_product_id,
):
    """직접경비 행의 매출구분 카운트.

    = (차량번호 ↔ 신번호 카운트)
    + (차량번호 ↔ 구번호 카운트)
    + (상품아이디 ↔ 상품ID 카운트)
    """
    car_key = _normalize_lookup_value(car_number)
    product_key = _normalize_lookup_value(source_product_id)
    counts = lookup[sales_type]
    return (
        counts["신번호"].get(car_key, 0)
        + counts["구번호"].get(car_key, 0)
        + counts["상품ID"].get(product_key, 0)
    )


def _build_segment_to_product_id_lookups(detail_df):
    """detail_df (product_id_df) 에서 다음 두 dict 반환.

    new_lookup: {매출구분}_{신번호} → 상품ID
    old_lookup: {매출구분}_{구번호} → 상품ID
    """
    new_lookup = {}
    old_lookup = {}
    if detail_df is None or detail_df.empty:
        return new_lookup, old_lookup

    detail = _strip_columns(detail_df)
    if "매출구분" not in detail.columns or "상품ID" not in detail.columns:
        return new_lookup, old_lookup

    detail = detail.copy()
    detail["_매출구분"] = detail["매출구분"].astype(str).str.strip()
    detail["_상품ID"] = detail["상품ID"].astype(str).str.strip()
    detail = detail[detail["_매출구분"].ne("") & detail["_상품ID"].ne("")]

    if "신번호" in detail.columns:
        _new = detail[["_매출구분", "_상품ID", "신번호"]].copy()
        _new["_신번호"] = _new["신번호"].apply(_normalize_lookup_value)
        _new = _new[_new["_신번호"].ne("")]
        _new["_key"] = _new["_매출구분"] + "_" + _new["_신번호"]
        # setdefault(첫 값 우선): 중복 제거 후 첫 행만
        _new = _new.drop_duplicates(subset=["_key"], keep="first")
        new_lookup.update(dict(zip(_new["_key"], _new["_상품ID"])))

    if "구번호" in detail.columns:
        _old = detail[["_매출구분", "_상품ID", "구번호"]].copy()
        _old["_구번호"] = _old["구번호"].apply(_normalize_lookup_value)
        _old = _old[_old["_구번호"].ne("")]
        _old["_key"] = _old["_매출구분"] + "_" + _old["_구번호"]
        _old = _old.drop_duplicates(subset=["_key"], keep="first")
        old_lookup.update(dict(zip(_old["_key"], _old["_상품ID"])))

    return new_lookup, old_lookup


def _build_sales_type_product_id_lookup(detail_df):
    """detail_df 에서 {매출구분}_{상품ID} → 상품ID lookup 반환."""
    lookup = {}
    if detail_df is None or detail_df.empty:
        return lookup

    detail = _strip_columns(detail_df)
    if "매출구분" not in detail.columns or "상품ID" not in detail.columns:
        return lookup

    detail = detail.copy()
    detail["_매출구분"] = detail["매출구분"].astype(str).str.strip()
    detail["_상품ID"] = detail["상품ID"].apply(_normalize_lookup_value)
    detail = detail[detail["_매출구분"].ne("") & detail["_상품ID"].ne("")]

    _tmp = detail[["_매출구분", "_상품ID"]].copy()
    _tmp["_key"] = _tmp["_매출구분"] + "_" + _tmp["_상품ID"]
    _tmp = _tmp.drop_duplicates(subset=["_key"], keep="first")
    lookup.update(dict(zip(_tmp["_key"], _tmp["_상품ID"])))

    return lookup


def _build_product_id_to_sales_type_lookup(detail_df):
    """detail_df 에서 상품ID → 매출구분 lookup 반환."""
    lookup = {}
    if detail_df is None or detail_df.empty:
        return lookup

    detail = _strip_columns(detail_df)
    if "매출구분" not in detail.columns or "상품ID" not in detail.columns:
        return lookup

    detail = detail.copy()
    detail["_매출구분"] = detail["매출구분"].astype(str).str.strip()
    detail["_상품ID"] = detail["상품ID"].apply(_normalize_lookup_value)
    detail = detail[detail["_매출구분"].ne("") & detail["_상품ID"].ne("")]

    _tmp = detail[["_상품ID", "_매출구분"]].drop_duplicates(subset=["_상품ID"], keep="first")
    lookup.update(dict(zip(_tmp["_상품ID"], _tmp["_매출구분"])))

    return lookup


def preprocess_material_cost_file(file, detail_df=None, settlement_year=None, settlement_month=None):
    """재료비"""
    sheets = _load_excel_sheets(file)
    sales_count_lookup = _build_vehicle_sales_count_lookup(detail_df)
    # 재료비 시트의 '구분자' (= 매출구분_차량번호) 로 상품ID 찾기 위한 lookup
    new_no_lookup, old_no_lookup = _build_segment_to_product_id_lookups(detail_df)

    processed_sheets = {}
    for sheet_name, df in sheets.items():
        if "출고부품분류" in df.columns:
            _cost_type = df["출고부품분류"].astype(str).str.strip()
            df["원가구분"] = np.select(
                [_cost_type.eq("원재료비 대체"), _cost_type.eq("페인트")],
                ["재료비", "페인트"],
                default="",
            )
        else:
            df["원가구분"] = ""

        # 확인용: 차량입고일자 기준 입고연도/입고월 (회계연도/회계월과 비교용)
        df = _set_accounting_year_month(df, "차량입고일자", year_col="입고연도", month_col="입고월")

        # 회계연도/회계월 필터
        if settlement_year is not None and "회계연도" in df.columns:
            df = df[pd.to_numeric(df["회계연도"], errors="coerce").eq(int(settlement_year))]
        if settlement_month is not None and "회계월" in df.columns:
            df = df[pd.to_numeric(df["회계월"], errors="coerce").eq(int(settlement_month))]
        if df.empty:
            continue

        # df.apply(axis=1) 대신 vectorized map으로 처리 (속도 개선)
        vehicle_keys = df["차량번호"].apply(
            lambda v: _normalize_lookup_value(v) if "차량번호" in df.columns else ""
        ) if "차량번호" in df.columns else pd.Series("", index=df.index)
        old_vehicle_keys = df["구차량번호"].apply(
            _normalize_lookup_value
        ) if "구차량번호" in df.columns else pd.Series("", index=df.index)

        for sales_type in MATERIAL_SALES_COLUMNS:
            new_map = sales_count_lookup[sales_type]["신번호"]
            old_map = sales_count_lookup[sales_type]["구번호"]
            df[sales_type] = (
                vehicle_keys.map(new_map).fillna(0).astype(int)
                + old_vehicle_keys.map(old_map).fillna(0).astype(int)
            )

        df["매출구분"] = np.select(
            [df["정비매출"].gt(0), df["위탁매출"].gt(0), df["사내매출"].gt(0)],
            ["정비매출", "위탁매출", "사내매출"],
            default="",
        )

        vehicle_numbers = (
            df["차량번호"].apply(lambda v: "" if pd.isna(v) else str(v).strip())
            if "차량번호" in df.columns
            else pd.Series([""] * len(df), index=df.index)
        )
        df["구분자"] = np.where(
            df["매출구분"].astype(str).str.strip().ne("") & vehicle_numbers.ne(""),
            df["매출구분"].astype(str).str.strip() + "_" + vehicle_numbers,
            "",
        )

        # 상품ID 매핑: 구분자(매출구분_신번호) → new_no_lookup, 없으면 old_no_lookup
        segment_keys = df["구분자"].apply(
            lambda v: str(v).strip() if pd.notna(v) else ""
        )
        df["상품ID"] = (
            segment_keys.map(new_no_lookup).fillna("")
            .where(segment_keys.map(new_no_lookup).notna() & segment_keys.map(new_no_lookup).ne(""),
                   segment_keys.map(old_no_lookup).fillna(""))
        )

        processed_sheets[sheet_name] = df

    return processed_sheets


# 배부대상 매핑 (적요 키워드 → 배부대상)
# 주의: 위에서부터 먼저 매칭됨(np.select). '정비판금파트' 는 '판금파트'/'정비파트' 를
#       부분 문자열로 포함하므로 반드시 그 두 규칙보다 위에 둬야 정비로 매칭된다.
_ALLOCATION_TARGET_RULES = [
    ("RQI", "RQI"),
    ("임차", "임차"),
    ("반납운영팀", "선물"),
    ("공정지원팀", "전체"),
    ("정비판금파트", "정비"),
    ("도장파트", "도장"),
    ("판금파트", "판금"),
    ("정비파트", "정비"),
]


def preprocess_combined_manufacturing_cost_files(
    files, cost_type, settlement_year=None, settlement_month=None,
):
    """노무비 / 부문별경비"""
    cost_frames = []
    for file in files:
        for df in _load_excel_sheets(file).values():
            if not df.empty:
                cost_frames.append(df)

    if not cost_frames:
        return pd.DataFrame()

    df = pd.concat(cost_frames, ignore_index=True)

    if "회계일자" in df.columns:
        accounting_date_text = df["회계일자"].astype(str).str.strip()
        df = df[~accounting_date_text.isin(["월계", "누계"])].copy()
    if "작성사원명" in df.columns:
        df = df[df["작성사원명"].astype(str).str.strip().ne("김겸윤")].copy()

    df["원가구분"] = cost_type
    df = _set_accounting_year_month(df, "회계일자")

    if settlement_year is not None and settlement_month is not None:
        df = df[
            (df["회계연도"] == int(settlement_year))
            & (df["회계월"] == int(settlement_month))
        ].copy()

    summary_text = (
        df["적요"].astype(str)
        if "적요" in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )
    conditions = [summary_text.str.contains(keyword, na=False) for keyword, _ in _ALLOCATION_TARGET_RULES]
    choices = [label for _, label in _ALLOCATION_TARGET_RULES]
    df["배부대상"] = np.select(conditions, choices, default="기타")

    return df.reset_index(drop=True)


def preprocess_direct_expense_file(file, detail_df=None):
    """직접경비"""
    sheets = _load_excel_sheets(file)
    sales_count_lookup = _build_direct_expense_sales_count_lookup(detail_df)
    new_no_lookup, old_no_lookup = _build_segment_to_product_id_lookups(detail_df)
    product_id_lookup = _build_sales_type_product_id_lookup(detail_df)
    sales_type_by_product_id = _build_product_id_to_sales_type_lookup(detail_df)

    processed_sheets = {}
    for sheet_name, df in sheets.items():
        # 원본 상품아이디/차량아이디 컬럼은 그대로 두고, 최종 매칭용 상품ID는 뒤에서 새로 만든다.
        source_product_id_column = next(
            (column for column in ["상품아이디", "차량아이디", "상품ID"] if column in df.columns),
            None,
        )
        df = _set_accounting_year_month(df, "매입일")

        vehicle_numbers = (
            df["차량번호"].apply(lambda v: "" if pd.isna(v) else str(v).strip())
            if "차량번호" in df.columns
            else pd.Series([""] * len(df), index=df.index)
        )
        source_product_ids = (
            df[source_product_id_column].apply(lambda v: "" if pd.isna(v) else str(v).strip())
            if source_product_id_column is not None
            else pd.Series([""] * len(df), index=df.index)
        )

        # 매출구분 카운트: 차량번호 ↔ 신번호/구번호 + 상품아이디 ↔ 상품ID 3개 합산
        for sales_type in MATERIAL_SALES_COLUMNS:
            df[sales_type] = [
                _get_direct_expense_sales_count(
                    sales_type, sales_count_lookup, car_number, product_id,
                )
                for car_number, product_id in zip(vehicle_numbers, source_product_ids)
            ]

        df["매출구분"] = np.select(
            [df["정비매출"].gt(0), df["위탁매출"].gt(0), df["사내매출"].gt(0)],
            ["정비매출", "위탁매출", "사내매출"],
            default="",
        )
        fallback_sales_type = source_product_ids.apply(
            lambda value: sales_type_by_product_id.get(_normalize_lookup_value(value), "")
        )
        df["매출구분"] = df["매출구분"].where(
            df["매출구분"].astype(str).str.strip().ne(""),
            fallback_sales_type,
        )

        sales_type_series = df["매출구분"].astype(str).str.strip()

        df["구분자1"] = np.where(
            sales_type_series.ne("") & vehicle_numbers.ne(""),
            sales_type_series + "_" + vehicle_numbers,
            "",
        )
        df["구분자2"] = np.where(
            sales_type_series.ne("") & source_product_ids.ne(""),
            sales_type_series + "_" + source_product_ids.apply(_normalize_lookup_value),
            "",
        )

        product_id_cache = {}
        product_ids = []
        for segment_key_1, segment_key_2 in zip(df["구분자1"], df["구분자2"]):
            key_1 = str(segment_key_1).strip() if pd.notna(segment_key_1) else ""
            key_2 = str(segment_key_2).strip() if pd.notna(segment_key_2) else ""
            cache_key = (key_1, key_2)
            if cache_key not in product_id_cache:
                product_id = ""
                if key_1:
                    product_id = new_no_lookup.get(key_1, "")
                    if not product_id:
                        product_id = old_no_lookup.get(key_1, "")
                if not product_id and key_2:
                    product_id = product_id_lookup.get(key_2, "")
                product_id_cache[cache_key] = product_id
            product_ids.append(product_id_cache[cache_key])
        df["상품ID"] = product_ids

        if "구분자2" in df.columns:
            ordered_columns = [column for column in df.columns if column != "상품ID"]
            insert_at = ordered_columns.index("구분자2") + 1
            ordered_columns = (
                ordered_columns[:insert_at] + ["상품ID"] + ordered_columns[insert_at:]
            )
            df = df[ordered_columns]

        processed_sheets[sheet_name] = df
    return processed_sheets