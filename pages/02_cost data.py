"""손익분석 (Cost Summary) - Streamlit 페이지

구성:
    1. 유틸 (DataFrame 표시용)
    2. 결산 연/월 선택
    3. 기초 DB 업로드
    4. 매입원가 업로드
    5. 제조원가 업로드
    6. 최종 원가 렌더링
    7. 페이지 엔트리
"""

import os
import re
import importlib

import numpy as np
import pandas as pd
import streamlit as st
from datetime import datetime

import cost_summary_preprocess as _cost_summary_preprocess
import cost_summary_builder as _cost_summary_builder

_cost_summary_preprocess = importlib.reload(_cost_summary_preprocess)
_cost_summary_builder = importlib.reload(_cost_summary_builder)

from cost_summary_preprocess import (
    BASE_DF_KEYS,
    HIDDEN_BASE_DF_KEYS,
    PRODUCT_ID_COLUMNS,
    build_product_id_detail_view,
    collect_product_ids,
    dataframe_for_display,
    dataframe_to_excel_bytes,
    filter_purchase_inquiry,
    workbook_to_excel_bytes,
    _excel_round,
    _build_detail_lookup,
    _build_direct_expense_sales_count_lookup,
    _build_product_id_to_sales_type_lookup,
    _build_sales_type_product_id_lookup,
    _build_segment_to_product_id_lookups,
    _build_vehicle_sales_count_lookup,
    _build_waste_resource_lookups,
    _prepare_consignment_ledger,
    _prepare_direct_expense_sheet,
    _prepare_manufacturing_cost_sheet,
    _prepare_material_cost_sheet,
    _prepare_opening_inventory,
    _prepare_payback_sheet,
    _prepare_product_ledger,
    _prepare_product_master,
    _prepare_purchase_inquiry,
    _prepare_sales,
    _prepare_waste_resource_sheet,
)
from cost_summary_builder import build_final_cost_df, get_last_manufacturing_expense_diagnostics, get_last_material_allocation_diagnostics


COST_SUMMARY_CALC_VERSION = "process-category-ts-v6"


def _build_input_fingerprint(*dfs_and_values):
    """빌드 입력의 지문 생성.

    DataFrame/dict 인자는 id() 기반으로 비교한다. 이 앱의 DataFrame/dict 들은 전부 자신만의
    _memoize 캐시 슬롯에서 나오므로(안 바뀌면 항상 같은 객체, 바뀌면 새 객체로 교체) id() 비교만으로
    변경 여부를 정확히 판단할 수 있고, 매 리런마다 전체 데이터를 훑는 비용을 없앨 수 있다.
    (예전에는 df.select_dtypes(...).sum().sum() 으로 전체 숫자 컬럼을 매 리런마다 스캔했음 — 리런이
    잦은 UPLOAD 탭에서 체감 지연/먹통의 주요 원인이었음)
    """
    parts = [("calc_version", COST_SUMMARY_CALC_VERSION)]
    for item in dfs_and_values:
        if isinstance(item, (pd.DataFrame, dict)):
            parts.append(("ref", id(item)))
        else:
            parts.append(("val", str(item)))
    return repr(parts)


def _files_fingerprint(files):
    """업로드 파일 리스트의 지문. 내용 대신 file_id/이름/크기만 사용(가볍고, 같은 파일이 재업로드되지 않는 한 충분히 유일함)."""
    if not files:
        return ()
    return tuple(
        (getattr(f, "file_id", None), getattr(f, "name", None), getattr(f, "size", None))
        for f in files
    )


def _master_file_state():
    """현재 마스터의 지문. 저장으로 마스터가 바뀌면 값이 달라져 캐시를 무효화하는 데 사용."""
    try:
        return _cost_summary_builder.master_state_fingerprint()
    except Exception:
        return None


def _memoize(cache_name, fingerprint_parts, compute_fn):
    """fingerprint_parts 가 이전 호출과 같으면 session_state 에 저장된 결과를 재사용.

    Streamlit 은 위젯 조작마다 스크립트 전체를 재실행하므로, 관련 없는 위젯을 건드렸을 때도
    업로드 파일을 매번 재파싱하지 않도록 하기 위함. 실제 파일/입력이 바뀌면 정상적으로 재계산됨.
    """
    cache_key = f"_memo_{cache_name}"
    fingerprint = _build_input_fingerprint(*fingerprint_parts)
    cached = st.session_state.get(cache_key)
    if cached is not None and cached.get("fingerprint") == fingerprint:
        return cached["result"]
    result = compute_fn()
    st.session_state[cache_key] = {"fingerprint": fingerprint, "result": result}
    return result


def _cached_build_final_cost_df(
    product_id_df, purchase_cost_sheet_dfs, inventory_df, settlement_month,
    manufacturing_cost_sheet_dfs, verification_sheets, settlement_year,
    cost_driver_dfs, combined_cost_driver_df, consignment_ledger_df,
):
    """입력 지문이 같으면 session_state 에 저장된 직전 빌드 결과를 재사용."""
    fingerprint = _build_input_fingerprint(
        product_id_df, purchase_cost_sheet_dfs, inventory_df, settlement_month,
        manufacturing_cost_sheet_dfs, verification_sheets, settlement_year,
        cost_driver_dfs, combined_cost_driver_df, consignment_ledger_df,
    )
    cache_key = "_final_cost_cache"
    cached = st.session_state.get(cache_key)
    if cached is not None and cached.get("fingerprint") == fingerprint:
        return cached["result"], cached.get("diagnostics", []), cached.get("material_diagnostics", [])

    result = build_final_cost_df(
        product_id_df,
        purchase_cost_sheet_dfs,
        inventory_df,
        settlement_month,
        manufacturing_cost_sheet_dfs=manufacturing_cost_sheet_dfs,
        verification_sheets=verification_sheets,
        settlement_year=settlement_year,
        cost_driver_dfs=cost_driver_dfs,
        combined_cost_driver_df=combined_cost_driver_df,
        consignment_ledger_df=consignment_ledger_df,
    )
    diagnostics = get_last_manufacturing_expense_diagnostics()
    material_diagnostics = result.attrs.get("재료비_배부내역") or get_last_material_allocation_diagnostics()
    st.session_state[cache_key] = {
        "fingerprint": fingerprint,
        "result": result,
        "diagnostics": diagnostics,
        "material_diagnostics": material_diagnostics,
    }
    return result, diagnostics, material_diagnostics


# ============================================================
# 1. 유틸
# ============================================================

def empty_product_id_df():
    return pd.DataFrame(columns=PRODUCT_ID_COLUMNS)


def initialize_base_dfs():
    return {key: None for key in BASE_DF_KEYS}


def render_dataframe_tabs(sheet_dfs):
    tabs = st.tabs(list(sheet_dfs.keys()))
    for i, sheet_name in enumerate(sheet_dfs.keys()):
        with tabs[i]:
            current_df = sheet_dfs[sheet_name]
            st.write(f"건수: {len(current_df):,}건")
            st.dataframe(dataframe_for_display(current_df), use_container_width=True)


def render_section_heading(text):
    """섹션 제목 왼쪽에 강조색 세로 바를 붙여 눈에 띄게 표시 (버튼과 같은 강조색 재사용)."""
    st.markdown(
        "<div style='border-left:4px solid #ff4b4b; line-height:1.15; "
        "padding:0 0 0 0.5rem; margin:0.4rem 0 0.2rem 0;'>"
        f"<span style='font-size:24px; font-weight:600; line-height:1.15;'>{text}</span>"
        "</div>",
        unsafe_allow_html=True,
    )


def render_sub_heading(text):
    """상위 섹션 제목(render_section_heading) 바로 아래 붙는 작은 소제목.

    위쪽은 섹션 제목과 살짝 띄우고, 아래쪽은 바로 뒤에 오는 업로드 위젯과 간격 없이 붙인다.
    """
    st.markdown(
        "<div style='margin-top:0.6rem; margin-bottom:-0.1rem;'>"
        f"<span style='font-size:20px; font-weight:600;'>{text}</span>"
        "</div>",
        unsafe_allow_html=True,
    )


def render_sheet_workbook(sheet_dfs, download_label, file_name, empty_message):
    if not sheet_dfs:
        st.info(empty_message)
        return
    st.download_button(
        download_label,
        # sheet_dfs 가 안 바뀐 채(id로 식별) 리런될 때 워크북 재생성 방지
        data=_memoize(
            f"workbook_download_{file_name}", (id(sheet_dfs),),
            lambda: workbook_to_excel_bytes(sheet_dfs),
        ),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    render_dataframe_tabs(sheet_dfs)


# ============================================================
# 2. 결산 연/월 선택
# ============================================================

def render_settlement_selector():
    if "settlement_year" not in st.session_state:
        st.session_state["settlement_year"] = pd.Timestamp.today().year
    if "settlement_month" not in st.session_state:
        st.session_state["settlement_month"] = pd.Timestamp.today().month

    render_section_heading("원가산출 기준월 지정")

    year_col, month_col, _spacer_col, button_col = st.columns([1, 1, 1, 0.6])

    with year_col:
        selected_year = st.number_input(
            "기준연도", min_value=2000, max_value=2100,
            value=int(st.session_state["settlement_year"]),
            step=1, key="selected_settlement_year",
        )
    with month_col:
        selected_month = st.selectbox(
            "기준월", options=list(range(1, 13)),
            index=st.session_state["settlement_month"] - 1,
            format_func=lambda month: f"{month}월",
            key="selected_settlement_month",
        )
    with button_col:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        apply_period = st.button(
            "기준연/월 적용", key="apply_settlement_period",
            use_container_width=True, type="primary",
        )

    if apply_period:
        st.session_state["settlement_year"] = int(selected_year)
        st.session_state["settlement_month"] = selected_month
        st.success(f"기준연/월이 {int(selected_year)}년 {selected_month}월로 지정되었습니다.")

    settlement_year = st.session_state["settlement_year"]
    settlement_month = st.session_state["settlement_month"]
    st.caption("기준연도/월 기준부터 원가가 재계산됩니다.")
    st.divider()
    return settlement_year, settlement_month


# ============================================================
# 3. 기초 DB 업로드
# ============================================================

def process_base_sheets(sheets, dfs, settlement_year, settlement_month):
    """시트명에 따라 기초 DB 전처리 라우팅 (기초재고는 매입조회가 먼저 필요해서 별도 처리)."""
    if "매입조회" in sheets:
        dfs["매입조회"] = _prepare_purchase_inquiry(
            sheets["매입조회"], settlement_year, settlement_month,
        )
    if "전체상품관리_선매입" in sheets:
        dfs["전체상품조회"] = _prepare_product_master(
            sheets["전체상품관리_선매입"], settlement_year, settlement_month,
        )
    if "위탁수불부" in sheets:
        ledger_all, ledger_opening, ledger_inbound = _prepare_consignment_ledger(
            sheets["위탁수불부"], settlement_year, settlement_month,
        )
        dfs["위탁수불부"] = ledger_all
        dfs["위탁수불부_전체"] = ledger_all
        dfs["위탁수불부_기초"] = ledger_opening
        dfs["위탁수불부_입고"] = ledger_inbound
    if "복수비용_검사" in sheets:
        dfs["검사매출"] = _prepare_sales(
            sheets["복수비용_검사"],
            settlement_year=settlement_year, settlement_month=settlement_month,
        )
    if "복수비용_정비" in sheets:
        dfs["정비매출"] = _prepare_sales(
            sheets["복수비용_정비"], exclude_partner="현대자동차(주)",
            settlement_year=settlement_year, settlement_month=settlement_month,
        )


def process_opening_inventory_sheet(sheets, dfs):
    """'기초재고' 시트 처리 (매입조회 필터링 포함)."""
    if "기초재고" not in sheets:
        return
    try:
        inventory_all, inventory_filtered = _prepare_opening_inventory(
            sheets["기초재고"], dfs["매입조회"]
        )
        dfs["기초재고_전체"] = inventory_all
        dfs["기초재고"] = inventory_filtered
        dfs["매입조회"] = filter_purchase_inquiry(dfs["매입조회"], dfs["기초재고_전체"])
    except Exception as exc:
        st.error(f"기초재고 시트 처리 중 오류: {exc}")


def compute_base_dfs_from_sheets(sheets, settlement_year, settlement_month, quiet=False):
    """원가대상 원본 시트 dict → 처리된 dfs (렌더링과 분리된 순수 계산).

    파일 업로드 위젯과 무관하게, 이미 읽어들인 시트 dict 와 결산연/월만으로 호출 가능
    (일괄 처리 시 같은 시트를 여러 결산월에 대해 반복 호출하기 위함).

    quiet=True 면 자동 생성/제외 안내(st.info/st.warning)를 표시하지 않는다 —
    여러 달을 한 번에 처리할 때 화면에 메시지가 줄줄이 쌓이는 걸 막기 위함
    (에러는 quiet 여부와 무관하게 항상 표시).
    """
    local_dfs = initialize_base_dfs()

    if sheets:
        try:
            process_base_sheets(sheets, local_dfs, settlement_year, settlement_month)
        except Exception as exc:
            st.error(f"원가대상 처리 중 오류: {exc}")

        process_opening_inventory_sheet(sheets, local_dfs)

    # 기초재고가 비어있으면 누적 마스터에서 자동 생성 (직전월 기말_수량==1)
    inventory_now = local_dfs.get("기초재고")
    if inventory_now is None or (
        isinstance(inventory_now, pd.DataFrame) and inventory_now.empty
    ):
        try:
            from cost_summary_builder import (
                _build_inventory_df_from_master,
                _get_previous_master_prepaid_product_ids,
            )
            prepaid_product_ids = _get_previous_master_prepaid_product_ids(
                settlement_year, settlement_month,
            )
            if (
                prepaid_product_ids
                and isinstance(local_dfs.get("매입조회"), pd.DataFrame)
                and not local_dfs["매입조회"].empty
            ):
                before_count = len(local_dfs["매입조회"])
                prepaid_filter_df = pd.DataFrame({
                    "상품ID": sorted(prepaid_product_ids),
                    "선매입여부": ["선매입"] * len(prepaid_product_ids),
                })
                local_dfs["매입조회"] = filter_purchase_inquiry(
                    local_dfs["매입조회"], prepaid_filter_df,
                )
                removed_count = before_count - len(local_dfs["매입조회"])
                if removed_count > 0 and not quiet:
                    st.info(
                        f"전월 마스터 선매입 차량 {removed_count:,}건을 "
                        "매입조회에서 제외했습니다."
                    )

            auto_inv = _build_inventory_df_from_master(
                settlement_year, settlement_month,
            )
            if auto_inv is not None and not auto_inv.empty:
                local_dfs["기초재고"] = auto_inv
                local_dfs["기초재고_전체"] = auto_inv
                prev_y = settlement_year - 1 if settlement_month == 1 else settlement_year
                prev_m = 12 if settlement_month == 1 else settlement_month - 1
                if not quiet:
                    st.info(
                        f"기초재고 파일이 업로드되지 않아 누적 마스터에서 "
                        f"{prev_y}-{prev_m:02d} 기말 데이터를 자동으로 가져왔습니다. "
                        f"({len(auto_inv):,}건)"
                    )
        except Exception as exc:
            st.warning(f"기초재고 자동 생성 중 오류: {exc}")

    return local_dfs


# 원가대상 시트별로 기간을 식별하는 (연도컬럼, 월컬럼) 후보
_PERIOD_DETECTION_SHEET_SPECS = [
    ("매입조회", "매입연도", "매입월"),
    ("전체상품관리_선매입", "매입연도", "매입월"),
    ("위탁수불부", "회계연도", "회계월"),
    ("복수비용_검사", "매입연도", "매입월"),
    ("복수비용_정비", "매입연도", "매입월"),
]


def detect_available_periods(base_sheets):
    """원가대상 원본 시트에서 실제 존재하는 (연도, 월) 조합을 모아 오름차순으로 반환.

    일괄 처리 모드에서 순차로 돌릴 결산월 목록을 만드는 데 사용.
    """
    if not base_sheets:
        return []

    periods = set()
    for sheet_name, year_col, month_col in _PERIOD_DETECTION_SHEET_SPECS:
        df = base_sheets.get(sheet_name)
        if df is None or df.empty:
            continue
        work = df.copy()
        work.columns = [str(c).strip() for c in work.columns]
        if year_col not in work.columns or month_col not in work.columns:
            continue
        years = pd.to_numeric(work[year_col], errors="coerce")
        months = pd.to_numeric(work[month_col], errors="coerce")
        valid = years.notna() & months.notna()
        for y, m in zip(years[valid], months[valid]):
            periods.add((int(y), int(m)))

    return sorted(periods)


def render_base_upload(settlement_year, settlement_month):
    render_section_heading("원가대상")

    uploaded_file = st.file_uploader(
        "업로드 파일 ㅣ cost data_대상 l 기초재고/ 매입조회/ 전체상품조회/ 위탁수불부/ 복수비용_검사/ 복수비용_정비 (총 6개 시트 포함)",
        type=["xlsx"],
    )

    def _read_base_sheets():
        if uploaded_file is None:
            return {}
        try:
            return {
                str(name).strip(): sheet_df
                for name, sheet_df in pd.read_excel(uploaded_file, sheet_name=None).items()
            }
        except Exception as exc:
            st.error(f"{uploaded_file.name} 처리 중 오류: {exc}")
            return {}

    base_sheets = _memoize(
        "base_upload_sheets",
        (_files_fingerprint([uploaded_file] if uploaded_file is not None else []),),
        _read_base_sheets,
    )

    dfs = _memoize(
        "base_upload",
        (
            _files_fingerprint([uploaded_file] if uploaded_file is not None else []),
            settlement_year, settlement_month,
            _master_file_state(),
        ),
        lambda: compute_base_dfs_from_sheets(base_sheets, settlement_year, settlement_month),
    )
    product_id_df = empty_product_id_df()

    if uploaded_file is not None or (
        isinstance(dfs.get("기초재고"), pd.DataFrame) and not dfs["기초재고"].empty
    ):
        st.divider()
        st.subheader("- 상품ID 확인")
        # dfs 가 안 바뀐 채(id로 식별) 리런될 때 재계산 방지 + product_id_df 자체도 안정적인 참조가 되어
        # 이 값을 지문에 사용하는 다른 캐시(_cached_build_final_cost_df 등)도 같이 효과를 봄
        product_id_df = _memoize(
            "collect_product_ids", (id(dfs), settlement_year, settlement_month),
            lambda: collect_product_ids(dfs, settlement_year, settlement_month),
        )

        if not product_id_df.empty:
            st.write(f"데이터 건수: {len(product_id_df):,}건")
            with st.expander("구분 포함 상세 보기"):
                product_id_detail_df = _memoize(
                    "product_id_detail_view", (id(product_id_df),),
                    lambda: build_product_id_detail_view(product_id_df),
                )
                st.download_button(
                    "엑셀 다운로드",
                    data=_memoize(
                        "product_id_detail_download", (id(product_id_detail_df),),
                        lambda: dataframe_to_excel_bytes(product_id_detail_df, sheet_name="구분포함상세"),
                    ),
                    file_name="product_id_detail.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                st.dataframe(dataframe_for_display(product_id_detail_df), use_container_width=True)
        else:
            st.info("상품ID를 가진 업로드 데이터가 아직 없습니다.")

    # 원하는 탭 순서 (없는 키는 자동 건너뜀, 그 외 키는 뒤에 자연 순서로 붙음)
    _BASE_TAB_ORDER = [
        "기초재고", "매입조회", "전체상품조회", "검사매출", "정비매출",
        "위탁수불부", "위탁수불부_기초", "위탁수불부_입고",
    ]
    # "전체상품조회" 는 화면 라벨만 "선매입" 으로 표시 (내부 키는 그대로)
    _BASE_TAB_LABEL_OVERRIDE = {"전체상품조회": "선매입"}

    visible_dfs_raw = {
        key: value
        for key, value in dfs.items()
        if key not in HIDDEN_BASE_DF_KEYS and value is not None
    }
    # 지정 순서 → 그 외 잔여 키 순서
    ordered_keys = [k for k in _BASE_TAB_ORDER if k in visible_dfs_raw]
    ordered_keys += [k for k in visible_dfs_raw if k not in ordered_keys]
    visible_dfs = {
        _BASE_TAB_LABEL_OVERRIDE.get(k, k): visible_dfs_raw[k]
        for k in ordered_keys
    }
    if visible_dfs:
        with st.expander("파일별 개별 데이터 확인"):
            render_dataframe_tabs(visible_dfs)

    return dfs, product_id_df, base_sheets


# ============================================================
# 4. 매입원가 업로드
# ============================================================

def process_purchase_cost_sheets(sheets, product_id_df, dfs, settlement_year, settlement_month):
    """시트명(상품원장/폐자원공제/페이백)에 따라 매입원가 전처리 라우팅."""
    cost_sheet_dfs = {}

    if "상품원장" in sheets:
        try:
            cost_sheet_dfs["상품원장"] = _prepare_product_ledger(
                sheets["상품원장"], product_id_df, dfs.get("기초재고_전체"),
                settlement_year, settlement_month,
                product_id_df=product_id_df,
            )
        except Exception as exc:
            st.error(f"상품원장 시트 처리 중 오류: {exc}")

    if "폐자원공제" in sheets:
        if product_id_df.empty:
            st.warning("폐자원공제 시트의 상품ID를 가져오려면 원가대상을 먼저 업로드하세요.")
        try:
            lookups = _build_waste_resource_lookups(None, product_id_df)
            cost_sheet_dfs["폐자원공제"] = _prepare_waste_resource_sheet(
                sheets["폐자원공제"], *lookups,
            )
        except Exception as exc:
            st.error(f"폐자원공제 시트 처리 중 오류: {exc}")

    if "페이백" in sheets:
        if product_id_df.empty:
            st.warning("페이백 시트의 상품ID를 가져오려면 원가대상도 함께 업로드하세요.")
        try:
            detail_lookup = _build_detail_lookup(product_id_df)
            cost_sheet_dfs["페이백"] = _prepare_payback_sheet(
                sheets["페이백"], detail_lookup, settlement_year, settlement_month,
            )
        except Exception as exc:
            st.error(f"페이백 시트 처리 중 오류: {exc}")

    return cost_sheet_dfs


def render_purchase_cost_upload(product_id_df, dfs, settlement_year, settlement_month):
    render_sub_heading("매입원가")

    uploaded_cost_file = st.file_uploader(
        "업로드 파일 ㅣ cost data_매입원가 l 상품원장/ 재활용폐자원세액공제신고서/ 페이백 (총 3개 시트 포함)",
        type=["xlsx", "xls"], key="cost_file",
    )

    if uploaded_cost_file is None:
        return {}, {}

    def _read_cost_sheets():
        try:
            return {
                str(name).strip(): sheet_df
                for name, sheet_df in pd.read_excel(uploaded_cost_file, sheet_name=None).items()
            }
        except Exception as exc:
            st.error(f"{uploaded_cost_file.name} 처리 중 오류: {exc}")
            return {}

    cost_sheets = _memoize(
        "purchase_cost_upload_sheets",
        (_files_fingerprint([uploaded_cost_file]),),
        _read_cost_sheets,
    )

    cost_sheet_dfs = _memoize(
        "purchase_cost_upload",
        (
            _files_fingerprint([uploaded_cost_file]), product_id_df, dfs,
            settlement_year, settlement_month,
        ),
        lambda: process_purchase_cost_sheets(
            cost_sheets, product_id_df, dfs, settlement_year, settlement_month,
        ),
    )
    render_sheet_workbook(
        cost_sheet_dfs,
        "매입원가 파일 다운로드",
        "purchase_cost_preprocessed.xlsx",
        "매입원가 파일에서 표시할 데이터가 없습니다.",
    )
    return cost_sheet_dfs, cost_sheets


# ============================================================
# 5. 제조원가 업로드
# ============================================================

def process_manufacturing_cost_sheets(sheets, product_id_df, settlement_year, settlement_month):
    """시트명(재료비/노무비/부문별경비/직접경비)에 따라 제조원가 전처리 라우팅."""
    cost_sheet_dfs = {}

    if "재료비" in sheets:
        try:
            sales_count_lookup = _build_vehicle_sales_count_lookup(product_id_df)
            new_no_lookup, old_no_lookup = _build_segment_to_product_id_lookups(product_id_df)
            result = _prepare_material_cost_sheet(
                sheets["재료비"], sales_count_lookup, new_no_lookup, old_no_lookup,
                settlement_year, settlement_month,
            )
            if result is not None:
                cost_sheet_dfs["재료비"] = result
        except Exception as exc:
            st.error(f"재료비 시트 처리 중 오류: {exc}")

    for sheet_name, cost_type in (("노무비", "노무비"), ("부문별경비", "제조경비")):
        if sheet_name not in sheets:
            continue
        try:
            cost_sheet_dfs[sheet_name] = _prepare_manufacturing_cost_sheet(
                sheets[sheet_name], cost_type, settlement_year, settlement_month,
            )
        except Exception as exc:
            st.error(f"{sheet_name} 시트 처리 중 오류: {exc}")

    if "직접경비" in sheets:
        try:
            sales_count_lookup = _build_direct_expense_sales_count_lookup(product_id_df)
            new_no_lookup, old_no_lookup = _build_segment_to_product_id_lookups(product_id_df)
            product_id_lookup = _build_sales_type_product_id_lookup(product_id_df)
            sales_type_by_product_id = _build_product_id_to_sales_type_lookup(product_id_df)
            cost_sheet_dfs["직접경비"] = _prepare_direct_expense_sheet(
                sheets["직접경비"], sales_count_lookup, new_no_lookup, old_no_lookup,
                product_id_lookup, sales_type_by_product_id,
            )
        except Exception as exc:
            st.error(f"직접경비 시트 처리 중 오류: {exc}")

    # 원하는 탭 순서: 재료비 → 노무비 → 부문별경비 → 직접경비
    _MFG_TAB_ORDER = ["재료비", "노무비", "부문별경비", "직접경비"]
    ordered_keys = [k for k in _MFG_TAB_ORDER if k in cost_sheet_dfs]
    ordered_keys += [k for k in cost_sheet_dfs if k not in ordered_keys]
    return {k: cost_sheet_dfs[k] for k in ordered_keys}


def render_manufacturing_cost_upload(product_id_df, settlement_year, settlement_month):
    render_sub_heading("제조원가")

    uploaded_file = st.file_uploader(
        "업로드 파일 ㅣ cost data_제조원가 l 재료비/ 노무비/ 부문별경비/ 직접경비 시트 포함 (총 4개 시트 포함)",
        type=["xlsx", "xls"], key="manufacturing_cost_file",
    )

    if uploaded_file is None:
        return {}, {}

    def _read_manufacturing_sheets():
        try:
            return {
                str(name).strip(): sheet_df
                for name, sheet_df in pd.read_excel(uploaded_file, sheet_name=None).items()
            }
        except Exception as exc:
            st.error(f"{uploaded_file.name} 처리 중 오류: {exc}")
            return {}

    manufacturing_sheets = _memoize(
        "manufacturing_cost_upload_sheets",
        (_files_fingerprint([uploaded_file]),),
        _read_manufacturing_sheets,
    )

    manufacturing_cost_sheet_dfs = _memoize(
        "manufacturing_cost_upload",
        (_files_fingerprint([uploaded_file]), product_id_df, settlement_year, settlement_month),
        lambda: process_manufacturing_cost_sheets(
            manufacturing_sheets, product_id_df, settlement_year, settlement_month,
        ),
    )

    render_sheet_workbook(
        manufacturing_cost_sheet_dfs,
        "제조원가 파일 다운로드",
        "manufacturing_cost_preprocessed.xlsx",
        "제조원가 파일에서 표시할 데이터가 없습니다.",
    )
    return manufacturing_cost_sheet_dfs, manufacturing_sheets


# ============================================================
# 6. 원가동인 업로드
# ============================================================

def _clean_cost_driver_sheet(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
    if df.empty:
        return df
    # '상품아이디' 컬럼을 '상품ID' 로 통일
    if "상품ID" not in df.columns and "상품아이디" in df.columns:
        df = df.rename(columns={"상품아이디": "상품ID"})
    return df


def process_cost_driver_sheets(sheets):
    """원가동인 원본 시트 dict → {keyword: {sheet_name: df}} (렌더링과 분리된 순수 계산).

    결산연/월과 무관 — 실제 결산기간 필터는 build_combined_cost_driver_df /
    _aggregate_cost_driver_by_product_id 단계에서 회계연도·회계월(또는 연도·월) 로 이뤄진다.
    """
    local_cost_driver_dfs = {}
    if not sheets:
        return local_cost_driver_dfs

    for keyword in ("RQI", "RTLS", "TS"):
        if keyword not in sheets:
            continue
        df = _clean_cost_driver_sheet(sheets[keyword])
        if not df.empty:
            local_cost_driver_dfs[keyword] = {keyword: df}

    if "RTC_SM" in sheets:
        df = _clean_cost_driver_sheet(sheets["RTC_SM"])
        if not df.empty:
            if "데이터구분" not in df.columns:
                st.warning("⚠️ RTC_SM 시트: '데이터구분' 컬럼이 없어 rtc/sm 분리 불가.")
            else:
                kind = df["데이터구분"].astype(str).str.strip().str.lower()
                rtc_df = df[kind.eq("rtc")].reset_index(drop=True)
                sm_df = df[kind.eq("sm")].reset_index(drop=True)
                if not rtc_df.empty:
                    local_cost_driver_dfs["rtc"] = {"RTC_SM": rtc_df}
                if not sm_df.empty:
                    local_cost_driver_dfs["sm"] = {"RTC_SM": sm_df}

    return local_cost_driver_dfs


def render_cost_driver_upload(settlement_year=None, settlement_month=None):
    render_section_heading("배부 기준자료")
    render_sub_heading("원가동인")

    uploaded_file = st.file_uploader(
        "업로드 파일 ㅣcost data_원가동인 l RQI/ TS/ RTLS/ RTC/ SM",
        type=["xlsx", "xls"], key="cost_driver_file",
    )

    if uploaded_file is None:
        return {}, {}

    def _read_driver_sheets():
        try:
            return {
                str(name).strip(): sheet_df
                for name, sheet_df in pd.read_excel(uploaded_file, sheet_name=None).items()
            }
        except Exception as exc:
            st.error(f"{uploaded_file.name} 처리 중 오류: {exc}")
            return {}

    driver_sheets = _memoize(
        "cost_driver_upload_sheets",
        (_files_fingerprint([uploaded_file]),),
        _read_driver_sheets,
    )

    cost_driver_dfs = _memoize(
        "cost_driver_upload",
        (_files_fingerprint([uploaded_file]),),
        lambda: process_cost_driver_sheets(driver_sheets),
    )

    if cost_driver_dfs:
        with st.expander("원가동인 데이터 확인 (선택된 시트만 표시)"):
            flattened = {
                f"{keyword} / {sheet_name}": df
                for keyword, sheet_map in cost_driver_dfs.items()
                for sheet_name, df in sheet_map.items()
            }
            render_dataframe_tabs(flattened)

    return cost_driver_dfs, driver_sheets


# ============================================================
# 원가동인 통합 (RQI / TS / RTLS → 공정 통합 DataFrame)
# ============================================================

# 통합 DataFrame 컬럼 순서
COMBINED_DRIVER_COLUMNS = [
    "공정", "차량번호", "모델명",
    "최초측정일", "최초측정시간", "최종측정일", "최종측정시간",
    "담당자",
]

# 구분 변환 규칙 (공정 → 구분)
# 엑셀 수식 동등:
#   IF(공정 IN {"RQI","리본카옥션성능","법적성능","TS"}, "RQI",
#     IF(공정="TU","정비",
#       IF(공정="PL","판금",
#         IF(공정="PA","도장", 공정))))
PROCESS_TO_CATEGORY = {
    "RQI": "RQI",
    "리본카옥션성능": "RQI",
    "차옥션성능": "RQI",
    "법적성능": "RQI",
    "TS": "RQI",
    "TU": "정비",
    "PL": "판금",
    "PA": "도장",
}


def _pick_first_existing(df, candidates, default=""):
    """후보 컬럼 중 가장 먼저 존재하는 것의 Series 반환. 없으면 default."""
    for column in candidates:
        if column in df.columns:
            return df[column]
    return pd.Series([default] * len(df), index=df.index)


def _normalize_rtls_sheet(df, sheet_name):
    """RTLS 시트 → 통합 컬럼 매핑 (공정은 원본 '공정' 컬럼 값)."""
    process_value = _pick_first_existing(df, ["공정"], default=sheet_name)
    return pd.DataFrame({
        "공정": process_value,
        "차량번호": _pick_first_existing(df, ["차량번호"]),
        "모델명": _pick_first_existing(df, ["차량모델", "모델명"]),
        "최초측정일": _pick_first_existing(df, ["최초측정일"]),
        "최초측정시간": _pick_first_existing(df, ["최초측정시간"]),
        "최종측정일": _pick_first_existing(df, ["최종측정일"]),
        "최종측정시간": _pick_first_existing(df, ["최종측정시간"]),
        "담당자": _pick_first_existing(df, ["작업자", "담당자"]),
    })


def _normalize_ts_sheet(df, sheet_name):
    """TS 시트 → 통합 컬럼 매핑 (공정은 'TS' 고정).
    TS 컬럼: 차량번호, 차량명, 검사일(20260102), (빈값), 검사일, (빈값), 검사자
    """
    inspection_date = _pick_first_existing(df, ["검사일"])
    return pd.DataFrame({
        "공정": "TS",
        "차량번호": _pick_first_existing(df, ["차량번호"]),
        "모델명": _pick_first_existing(df, ["차량명", "모델명"]),
        "최초측정일": inspection_date,
        "최초측정시간": "",
        "최종측정일": inspection_date,
        "최종측정시간": "",
        "담당자": _pick_first_existing(df, ["검사자", "담당자"]),
    })


def _normalize_rqi_sheet(df, sheet_name):
    """RQI 시트 → 통합 컬럼 매핑 (구분 컬럼 값을 공정으로 사용, 없으면 공정 컬럼, 그것도 없으면 'RQI' 고정).
    컬럼: 차량번호, 차명, 작업시작일자, 작업시작시간, 작업종료일자, 작업종료시간, 담당자
    (구버전 파일은 시작/종료일자 구분 없이 '작업일자' 하나만 있어 fallback 으로 사용)
    """
    process_value = _pick_first_existing(df, ["구분", "공정"], default="RQI")
    return pd.DataFrame({
        "공정": process_value,
        "차량번호": _pick_first_existing(df, ["차량번호"]),
        "모델명": _pick_first_existing(df, ["차명", "모델명"]),
        "최초측정일": _pick_first_existing(df, ["작업시작일자", "작업일자"]),
        "최초측정시간": _pick_first_existing(df, ["작업시작시간"]),
        "최종측정일": _pick_first_existing(df, ["작업종료일자", "작업일자"]),
        "최종측정시간": _pick_first_existing(df, ["작업종료시간"]),
        "담당자": _pick_first_existing(df, ["담당자"]),
    })


def _normalize_process_to_category(process_value):
    """공정 값을 구분(정비/판금/도장)으로 변환. 매핑에 없으면 원값 그대로."""
    text = "" if pd.isna(process_value) else str(process_value).strip()
    return PROCESS_TO_CATEGORY.get(text, text)


# 영업시간 (일별)
WORK_START_HOUR = 8
WORK_START_MINUTE = 30
WORK_END_HOUR = 17
WORK_END_MINUTE = 30


def _parse_date_value(value):
    """다양한 형식의 날짜 → pandas Timestamp (실패 시 NaT).

    지원: datetime, '2026-01-02', '20260102' (8자리), 20260102 (정수) 등
    """
    if pd.isna(value) or value == "":
        return pd.NaT

    # 8자리 정수/문자 형태 ('20260102' 또는 20260102)
    text = str(value).strip()
    if len(text) == 8 and text.isdigit():
        return pd.to_datetime(text, format="%Y%m%d", errors="coerce")

    return pd.to_datetime(value, errors="coerce")



def _numeric_month_series(series):
    """'2', '02', '2월' 같은 월 값을 숫자로 변환."""
    text = series.astype(str).str.strip()
    text = text.str.replace(r"\.0$", "", regex=True)
    text = text.str.extract(r"(\d{1,2})", expand=False)
    return pd.to_numeric(text, errors="coerce")


def _filter_cost_driver_sheet_by_settlement_period(df, settlement_year, settlement_month):
    """원가동인 원본 행을 결산연도/월에 해당하는 행만 남긴다."""
    if (
        df is None or df.empty
        or settlement_year is None or settlement_month is None
    ):
        return df

    year = int(settlement_year)
    month = int(settlement_month)
    work = df.copy()

    if "회계연도" in work.columns and "회계월" in work.columns:
        year_values = pd.to_numeric(work["회계연도"], errors="coerce")
        month_values = _numeric_month_series(work["회계월"])
        return work[year_values.eq(year) & month_values.eq(month)].copy()

    return work


def _parse_time_value(value):
    """시간 문자열/객체 → datetime.time (실패 시 None).

    지원: '09:00', '09:00:00', datetime.time, datetime, pandas Timestamp 등
    """
    if pd.isna(value) or value == "":
        return None

    # 이미 time 객체
    if hasattr(value, "hour") and hasattr(value, "minute") and not hasattr(value, "year"):
        return value
    # datetime / Timestamp
    if hasattr(value, "hour") and hasattr(value, "year"):
        return value.time()

    text = str(value).strip()
    if not text:
        return None
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.notna(parsed):
        return parsed.time()
    return None


# 한국 공휴일 백업 (holidays 패키지 미설치 환경용)
# 필요 시 매년 갱신
KOREAN_HOLIDAYS_FALLBACK = {
    2025: [
        "2025-01-01", "2025-01-27", "2025-01-28", "2025-01-29", "2025-01-30",
        "2025-03-01", "2025-03-03", "2025-05-01", "2025-05-05", "2025-05-06", "2025-06-03",
        "2025-06-06", "2025-08-15", "2025-10-03", "2025-10-06", "2025-10-07",
        "2025-10-08", "2025-10-09", "2025-12-25",
    ],
    2026: [
        "2026-01-01", "2026-02-16", "2026-02-17", "2026-02-18", "2026-03-01",
        "2026-03-02", "2026-05-01", "2026-05-05", "2026-05-24", "2026-05-25", "2026-06-03",
        "2026-06-06", "2026-07-17", "2026-08-15", "2026-08-17", "2026-09-24",
        "2026-09-25", "2026-09-26", "2026-10-03", "2026-10-05", "2026-10-09",
        "2026-12-25",
    ],
    2027: [
        "2027-01-01", "2027-02-06", "2027-02-07", "2027-02-08", "2027-02-09",
        "2027-03-01", "2027-05-01", "2027-05-05", "2027-05-13", "2027-06-06", "2027-06-07",
        "2027-08-15", "2027-08-16", "2027-09-14", "2027-09-15", "2027-09-16",
        "2027-10-03", "2027-10-04", "2027-10-09", "2027-10-11", "2027-12-25",
    ],
}


@st.cache_data(show_spinner=False)
def _build_korean_holiday_checker(year_hint=None):
    """한국 공휴일 set (datetime.date 집합) 반환.

    1순위: holidays 패키지 사용
    2순위: KOREAN_HOLIDAYS_FALLBACK 하드코딩 데이터 사용
    (결산연도만 입력이라 캐싱 안전 — 매 리런마다 재계산 방지)
    """
    if year_hint is None:
        year_hint = pd.Timestamp.today().year
    year = int(year_hint)
    years = {year - 1, year, year + 1}

    # 1순위: holidays 패키지
    try:
        import holidays
        import datetime as _dt
        kr = holidays.country_holidays("KR", years=sorted(years))
        holiday_set = {d for d in kr.keys()}
        # 근로자의 날(5월 1일)은 holidays 패키지에 미포함 → 수동 추가
        for y in years:
            holiday_set.add(_dt.date(y, 5, 1))
        return holiday_set
    except Exception:
        pass

    # 2순위: 하드코딩 백업
    import datetime as _dt
    holiday_dates = set()
    for y in years:
        for date_str in KOREAN_HOLIDAYS_FALLBACK.get(y, []):
            try:
                holiday_dates.add(
                    _dt.datetime.strptime(date_str, "%Y-%m-%d").date()
                )
            except Exception:
                continue
    return holiday_dates


def _is_workday(date_obj, holiday_set):
    """평일이면서 공휴일이 아니면 True."""
    if pd.isna(date_obj):
        return False
    # 월=0 ... 일=6, 토(5)/일(6) 제외
    if date_obj.weekday() >= 5:
        return False
    if hasattr(date_obj, "date"):
        return date_obj.date() not in holiday_set
    return date_obj not in holiday_set


def _calculate_measurement_hours(
    start_date, start_time, end_date, end_time, holiday_set,
):
    """업무시간(08:30~17:30, 공휴일/주말 제외) 안에서 시작~종료 사이 시간(시간 단위) 계산."""
    sd = _parse_date_value(start_date)
    ed = _parse_date_value(end_date)
    st_time = _parse_time_value(start_time)
    et_time = _parse_time_value(end_time)

    if pd.isna(sd) or pd.isna(ed) or st_time is None or et_time is None:
        return ""

    work_start = pd.Timestamp.combine(
        sd.date(), pd.Timestamp(0).replace(
            hour=WORK_START_HOUR, minute=WORK_START_MINUTE,
        ).time(),
    )
    work_end = pd.Timestamp.combine(
        sd.date(), pd.Timestamp(0).replace(
            hour=WORK_END_HOUR, minute=WORK_END_MINUTE,
        ).time(),
    )

    start_dt = pd.Timestamp.combine(sd.date(), st_time)
    end_dt = pd.Timestamp.combine(ed.date(), et_time)
    if end_dt < start_dt:
        return 0.0

    total_seconds = 0.0
    current_day = sd.normalize()
    last_day = ed.normalize()
    # 안전장치 (최대 365일까지만)
    safety_counter = 0
    while current_day <= last_day and safety_counter < 366:
        safety_counter += 1
        day_start = pd.Timestamp.combine(
            current_day.date(),
            pd.Timestamp(0).replace(hour=WORK_START_HOUR, minute=WORK_START_MINUTE).time(),
        )
        day_end = pd.Timestamp.combine(
            current_day.date(),
            pd.Timestamp(0).replace(hour=WORK_END_HOUR, minute=WORK_END_MINUTE).time(),
        )

        # 영업일이 아니면 0
        if _is_workday(current_day, holiday_set):
            # 해당 일의 유효 구간: max(start, 08:30) ~ min(end, 17:30)
            segment_start = max(start_dt, day_start)
            segment_end = min(end_dt, day_end)
            if segment_end > segment_start:
                total_seconds += (segment_end - segment_start).total_seconds()

        current_day = current_day + pd.Timedelta(days=1)

    return _excel_round(total_seconds / 3600.0, 2)


def _build_segment_to_product_id_lookups_local(detail_df):
    """현재 페이지 모듈 내에서 사용하기 위해 preprocess 모듈 함수를 호출.

    detail_df 가 비어있으면 빈 dict 반환.
    """
    if detail_df is None or detail_df.empty:
        return {}, {}
    try:
        from cost_summary_preprocess import _build_segment_to_product_id_lookups
        return _build_segment_to_product_id_lookups(detail_df)
    except Exception:
        return {}, {}


def _enrich_with_sales_type_and_product_id(
    combined_df, product_id_df,
):
    """통합 DataFrame 에 정비매출/위탁매출/사내매출/매출구분/구분자/상품ID 컬럼 추가.

    재료비 시트와 동일 로직:
        - product_id_df 에서 (매출구분, 신번호) 또는 (매출구분, 구번호) 별 카운트로
          각 매출구분 컬럼 채움
        - 매출구분 우선순위: 정비매출 → 위탁매출 → 사내매출
        - 구분자 = '{매출구분}_{차량번호}'
        - 상품ID = (매출구분, 신번호) 또는 (매출구분, 구번호) 로 product_id_df 매칭
    """
    if combined_df.empty:
        for col in ["정비매출", "위탁매출", "사내매출", "매출구분", "구분자", "상품ID"]:
            combined_df[col] = "" if col in ("매출구분", "구분자", "상품ID") else 0
        return combined_df

    # 매출구분별 차량번호 카운트 lookup
    sales_count_lookup = {
        "정비매출": {"신번호": {}, "구번호": {}},
        "위탁매출": {"신번호": {}, "구번호": {}},
        "사내매출": {"신번호": {}, "구번호": {}},
    }
    if product_id_df is not None and not product_id_df.empty:
        if all(c in product_id_df.columns for c in ["매출구분", "신번호", "구번호"]):
            detail = product_id_df.copy()
            detail["_매출구분"] = detail["매출구분"].astype(str).str.strip()
            detail["_신번호"] = detail["신번호"].astype(str).str.strip().apply(
                lambda v: re.sub(r"\s+", "", v) if v else ""
            )
            detail["_구번호"] = detail["구번호"].astype(str).str.strip().apply(
                lambda v: re.sub(r"\s+", "", v) if v else ""
            )
            for sales_type in sales_count_lookup.keys():
                rows = detail[detail["_매출구분"].eq(sales_type)]
                sales_count_lookup[sales_type]["신번호"] = (
                    rows.loc[rows["_신번호"].ne(""), "_신번호"].value_counts().to_dict()
                )
                sales_count_lookup[sales_type]["구번호"] = (
                    rows.loc[rows["_구번호"].ne(""), "_구번호"].value_counts().to_dict()
                )

    new_no_lookup, old_no_lookup = _build_segment_to_product_id_lookups_local(product_id_df)

    # 각 행마다 매출구분별 카운트 계산
    car_keys = combined_df["차량번호"].apply(
        lambda v: re.sub(r"\s+", "", str(v).strip()) if pd.notna(v) and str(v).strip() else ""
    )

    for sales_type in ["정비매출", "위탁매출", "사내매출"]:
        combined_df[sales_type] = [
            sales_count_lookup[sales_type]["신번호"].get(key, 0)
            + sales_count_lookup[sales_type]["구번호"].get(key, 0)
            for key in car_keys
        ]

    # 매출구분 우선순위:
    #   1) 공정 == "TS" → 검사매출 (공정 자체로 결정)
    #   2) 정비매출 > 0 → 정비매출
    #   3) 위탁매출 > 0 → 위탁매출
    #   4) 사내매출 > 0 → 사내매출
    process_series = (
        combined_df["공정"].astype(str).str.strip()
        if "공정" in combined_df.columns
        else pd.Series([""] * len(combined_df), index=combined_df.index)
    )
    combined_df["매출구분"] = np.select(
        [
            process_series.eq("TS"),
            combined_df["정비매출"].gt(0),
            combined_df["위탁매출"].gt(0),
            combined_df["사내매출"].gt(0),
        ],
        ["검사매출", "정비매출", "위탁매출", "사내매출"],
        default="",
    )

    # 구분자 = 매출구분_차량번호
    sales_type_series = combined_df["매출구분"].astype(str).str.strip()
    vehicle_numbers = combined_df["차량번호"].apply(
        lambda v: "" if pd.isna(v) else str(v).strip()
    )
    combined_df["구분자"] = np.where(
        sales_type_series.ne("") & vehicle_numbers.ne(""),
        sales_type_series + "_" + vehicle_numbers,
        "",
    )

    # 상품ID = (매출구분, 신번호) 또는 (매출구분, 구번호) lookup
    product_ids = []
    cache = {}
    for segment_key in combined_df["구분자"]:
        key = str(segment_key).strip() if pd.notna(segment_key) else ""
        if not key:
            product_ids.append("")
            continue
        if key not in cache:
            pid = new_no_lookup.get(key, "")
            if not pid:
                pid = old_no_lookup.get(key, "")
            cache[key] = pid
        product_ids.append(cache[key])
    combined_df["상품ID"] = product_ids

    return combined_df


def build_combined_cost_driver_df(
    cost_driver_dfs,
    settlement_year=None,
    settlement_month=None,
    product_id_df=None,
):
    """원가동인 dict (RQI/TS/RTLS) 를 단일 DataFrame 으로 통합.

    반환 컬럼: [공정, 차량번호, 모델명, 최초측정일, 최초측정시간,
              최종측정일, 최종측정시간, 담당자, 구분, 측정시간(H),
              발생연도, 발생월,
              정비매출, 위탁매출, 사내매출, 매출구분, 구분자, 상품ID]
    """
    extra_columns = [
        "구분", "측정시간(H)", "발생연도", "발생월",
        "정비매출", "위탁매출", "사내매출", "매출구분", "구분자", "상품ID",
    ]
    if not cost_driver_dfs:
        return pd.DataFrame(columns=COMBINED_DRIVER_COLUMNS + extra_columns)

    normalizers = {
        "RTLS": _normalize_rtls_sheet,
        "TS": _normalize_ts_sheet,
        "RQI": _normalize_rqi_sheet,
    }

    frames = []
    for keyword, normalizer in normalizers.items():
        sheet_map = cost_driver_dfs.get(keyword)
        if not sheet_map:
            continue
        for sheet_name, df in sheet_map.items():
            if df is None or df.empty:
                continue
            try:
                df = _filter_cost_driver_sheet_by_settlement_period(
                    df, settlement_year, settlement_month,
                )
                if df is None or df.empty:
                    continue
                normalized = normalizer(df, sheet_name)
            except Exception as exc:
                st.warning(f"[{keyword} / {sheet_name}] 통합 중 오류: {exc}")
                continue
            if not normalized.empty:
                frames.append(normalized)

    if not frames:
        return pd.DataFrame(columns=COMBINED_DRIVER_COLUMNS + extra_columns)

    combined = pd.concat(frames, ignore_index=True)
    combined = combined[COMBINED_DRIVER_COLUMNS].copy()

    # 구분 컬럼: TU/PL/PA → 정비/판금/도장 (그 외는 원값)
    combined["구분"] = combined["공정"].apply(_normalize_process_to_category)

    # 측정시간(H) 계산
    holiday_set = _build_korean_holiday_checker(year_hint=settlement_year)

    def _row_hours(row):
        if str(row["공정"]).strip() == "TS":
            return 0.5
        return _calculate_measurement_hours(
            row["최초측정일"], row["최초측정시간"],
            row["최종측정일"], row["최종측정시간"],
            holiday_set,
        )
    combined["측정시간(H)"] = combined.apply(_row_hours, axis=1)

    # 발생연도/발생월 = 결산연도/결산월
    combined["발생연도"] = int(settlement_year) if settlement_year is not None else pd.NA
    combined["발생월"] = int(settlement_month) if settlement_month is not None else pd.NA

    # 매출구분/구분자/상품ID 채우기 (재료비와 동일 로직)
    combined = _enrich_with_sales_type_and_product_id(combined, product_id_df)

    # 최종 컬럼 순서
    return combined[COMBINED_DRIVER_COLUMNS + extra_columns]


def render_combined_cost_driver(
    cost_driver_dfs, settlement_year=None, settlement_month=None, product_id_df=None,
):
    """통합 원가동인 DataFrame 표시."""
    # cost_driver_dfs 는 시트별 df 를 담은 중첩 dict라 내용 기반 지문이 얕아질 수 있어,
    # (업스트림에서 파일이 안 바뀌면 같은 dict 객체를 그대로 재사용하는 점을 이용해) id() 로 식별.
    combined_df = _memoize(
        "combined_cost_driver",
        (id(cost_driver_dfs), product_id_df, settlement_year, settlement_month),
        lambda: build_combined_cost_driver_df(
            cost_driver_dfs, settlement_year, settlement_month, product_id_df,
        ),
    )
    if combined_df.empty:
        return combined_df

    with st.expander("원가동인 통합 (RQI / TS / RTLS)"):
        st.write(f"통합 건수: {len(combined_df):,}건")

        # 공휴일 처리 상태 표시 (측정시간 계산 디버깅용)
        if settlement_year is not None:
            holiday_set = _build_korean_holiday_checker(settlement_year)
            year_holidays = sorted(
                d for d in holiday_set if d.year == int(settlement_year)
            )
            try:
                import holidays as _h  # noqa: F401
                holiday_source = "holidays 패키지"
            except Exception:
                holiday_source = "내장 백업 데이터 (KOREAN_HOLIDAYS_FALLBACK)"
            st.caption(
                f"공휴일 처리: {holiday_source} | "
                f"{int(settlement_year)}년 공휴일 {len(year_holidays)}개"
            )

        st.download_button(
            "통합 원가동인 다운로드",
            data=_memoize(
                "combined_cost_driver_download", (id(combined_df),),
                lambda: dataframe_to_excel_bytes(combined_df, sheet_name="원가동인통합"),
            ),
            file_name="cost_driver_combined.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="cost_driver_combined_download",
        )
        st.dataframe(dataframe_for_display(combined_df), use_container_width=True)

    return combined_df


# ============================================================
# 7. 검증시트 업로드
# ============================================================

def render_verification_sheet_upload():
    render_sub_heading("기간별제조원가보고서")

    uploaded_file = st.file_uploader(
        "업로드 파일 ㅣ 총 1개 파일 ㅣ 검증용 ",
        type=["xlsx", "xls"], accept_multiple_files=False, key="verification_sheet_file",
    )

    if uploaded_file is None:
        return None

    def _compute_verification_sheets():
        try:
            # header=0 → 1번째 행이 컬럼명
            sheets = pd.read_excel(uploaded_file, sheet_name=None, header=0)
        except Exception as exc:
            st.error(f"{uploaded_file.name} 처리 중 오류: {exc}")
            return None

        # 시트가 여러 개일 수 있으니 dict 형태로 반환
        local_cleaned_sheets = {}
        for sheet_name, df in sheets.items():
            df.columns = [str(c).strip() for c in df.columns]
            df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
            local_cleaned_sheets[sheet_name] = df
        return local_cleaned_sheets

    cleaned_sheets = _memoize(
        "verification_sheet_upload",
        (_files_fingerprint([uploaded_file]),),
        _compute_verification_sheets,
    )

    if cleaned_sheets:
        with st.expander("검증시트 데이터 확인"):
            render_dataframe_tabs(cleaned_sheets)

    return cleaned_sheets


# ============================================================
# 8. 최종 원가 렌더링
# ============================================================

def render_final_cost(
    product_id_df, purchase_cost_sheet_dfs, dfs, settlement_month,
    manufacturing_cost_sheet_dfs=None,
    verification_sheets=None,
    settlement_year=None,
    cost_driver_dfs=None,
    combined_cost_driver_df=None,
):
    render_section_heading("차량별 원가현황")

    final_cost_df, _cached_diag, _cached_material_diag = _cached_build_final_cost_df(
        product_id_df,
        purchase_cost_sheet_dfs,
        dfs.get("기초재고_전체"),
        settlement_month,
        manufacturing_cost_sheet_dfs,
        verification_sheets,
        settlement_year,
        cost_driver_dfs,
        combined_cost_driver_df,
        dfs.get("위탁수불부_전체"),
    )

    if final_cost_df.empty:
        st.info("원가대상 데이터를 업로드하면 최종 원가 초안이 생성됩니다.")
        return final_cost_df

    if not purchase_cost_sheet_dfs:
        st.info("매입원가의 상품원장 데이터를 업로드하면 금액 컬럼이 채워집니다.")

    # 재료비/노무비/제조경비 배부 내역 통합 expander
    material_diagnostics = _cached_material_diag or final_cost_df.attrs.get("재료비_배부내역") or get_last_material_allocation_diagnostics()
    expense_diagnostics = _cached_diag or final_cost_df.attrs.get("제조경비_배부내역") or get_last_manufacturing_expense_diagnostics()

    all_diagnostics = list(material_diagnostics or []) + list(expense_diagnostics or [])
    if all_diagnostics:
        with st.expander("🔍 재료비/노무비/제조경비 배부 내역"):
            diag_df = pd.DataFrame(all_diagnostics)
            display_diag = diag_df.copy()

            preferred_order = [
                "구분", "컬럼", "배부총액(분자)", "실제배부값합",
                "가중치합(분모)", "단가", "비고",
            ]
            ordered = [c for c in preferred_order if c in display_diag.columns]
            ordered += [c for c in display_diag.columns if c not in ordered]
            display_diag = display_diag[ordered]

            for col in ["배부총액(분자)", "실제배부값합"]:
                if col in display_diag.columns:
                    display_diag[col] = display_diag[col].apply(
                        lambda v: f"{_excel_round(v):,}" if isinstance(v, (int, float)) else v
                    )
            if "가중치합(분모)" in display_diag.columns:
                display_diag["가중치합(분모)"] = display_diag["가중치합(분모)"].apply(
                    lambda v: (f"{v:,.4f}".rstrip("0").rstrip(".") if isinstance(v, (int, float)) else v)
                )
            if "단가" in display_diag.columns:
                display_diag["단가"] = display_diag["단가"].apply(
                    lambda v: f"{v:,.4f}".rstrip("0").rstrip(".") if isinstance(v, (int, float)) else v
                )
            st.dataframe(display_diag, use_container_width=True)

    st.download_button(
        "차량별 원가 다운로드",
        data=_memoize(
            "final_cost_draft_download", (id(final_cost_df),),
            lambda: dataframe_to_excel_bytes(final_cost_df, sheet_name="차량별원가"),
        ),
        file_name="final_cost_draft.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.write(f"건수: {len(final_cost_df):,}건")
    st.dataframe(dataframe_for_display(final_cost_df), use_container_width=True)

    return final_cost_df


# ============================================================
# 9. 페이지 엔트리
# ============================================================

st.set_page_config(page_title="손익분석", layout="wide")
st.title("Cost Data")

# 최종 마스터 저장 (SQLite, cost_summary_builder.py 에 저장소 로직을 통합)
# 이 페이지에서는 얇은 위임 함수만 두고, 실제 저장/조회/삭제/락(트랜잭션) 로직은
# cost_summary_builder.py 쪽에 있다 — 그 모듈의 내부 조회 함수들(_get_previous_master_*,
# _build_inventory_df_from_master 등)도 같은 저장소를 읽어야 하므로 한 곳에 모아둔다.


def save_final_master(final_cost_df):
    """최종원가 결과를 누적해 서버에 저장 (공용). 반환: (저장 시각, 누적 후 총 행 수, 이번 빌드 행 수, 교체된 행 수)."""
    return _cost_summary_builder.save_final_master(final_cost_df)


def run_batch_periods(
    base_sheets, purchase_cost_sheets, manufacturing_sheets, cost_driver_sheets,
    verification_sheets, periods,
):
    """감지된 (연도, 월) 목록을 오름차순으로 순차 처리.

    각 결산월을 계산한 직후 바로 save_final_master() 로 저장한다 — 다음 결산월의
    기초재고 자동생성(_build_inventory_df_from_master)이 방금 저장한 기말재고를
    그대로 이어받아, 파일을 한 번만 올려도 1월→2월→3월... 순으로 이어서 계산되게 하기 위함.
    """
    cost_driver_dfs_raw = process_cost_driver_sheets(cost_driver_sheets)
    status_placeholder = st.empty()

    results = []
    for year, month in periods:
        label = f"{year}-{month:02d}"
        status_placeholder.info(f"{year}년 {month}월 계산중입니다.")
        try:
            local_dfs = compute_base_dfs_from_sheets(base_sheets, year, month, quiet=True)
            product_id_df = collect_product_ids(local_dfs, year, month)

            purchase_cost_sheet_dfs = process_purchase_cost_sheets(
                purchase_cost_sheets, product_id_df, local_dfs, year, month,
            )
            manufacturing_cost_sheet_dfs = process_manufacturing_cost_sheets(
                manufacturing_sheets, product_id_df, year, month,
            )
            combined_cost_driver_df = build_combined_cost_driver_df(
                cost_driver_dfs_raw, year, month, product_id_df,
            )

            final_cost_df, _diag, _material_diag = _cached_build_final_cost_df(
                product_id_df,
                purchase_cost_sheet_dfs,
                local_dfs.get("기초재고_전체"),
                month,
                manufacturing_cost_sheet_dfs,
                verification_sheets,
                year,
                cost_driver_dfs_raw,
                combined_cost_driver_df,
                local_dfs.get("위탁수불부_전체"),
            )

            if final_cost_df is None or final_cost_df.empty:
                results.append({"결산연월": label, "건수": 0, "상태": "데이터 없음 (건너뜀)"})
                continue

            saved_at, total_rows, new_rows, replaced_rows = save_final_master(final_cost_df)
            results.append({
                "결산연월": label,
                "건수": len(final_cost_df),
                "상태": f"저장 완료 (누적 {total_rows:,}건, 저장 {saved_at})",
            })
        except Exception as exc:
            results.append({"결산연월": label, "건수": 0, "상태": f"오류: {exc}"})

    status_placeholder.empty()
    return results


def load_final_master():
    """저장된 최종원가 마스터 불러오기. (df, 저장시각) 또는 (None, None)."""
    return _cost_summary_builder.read_final_master()


def delete_final_master():
    """마스터 데이터 전체 삭제. 삭제 성공 여부 반환."""
    return _cost_summary_builder.delete_final_master()


# ----- 회계처리 표 -----

# 월 + 누계 컬럼 (각 항목마다 대수/금액 2개씩)
_ACCOUNTING_MONTH_LABELS = [f"{m}월" for m in range(1, 13)] + ["누계"]

# 위탁매출 거래처 목록 (이 순서대로 표시, '기타'는 나머지)
_CONSIGNMENT_PARTNERS = [
    "하나캐피탈", "우리금융캐피탈", "NH농협캐피탈", "삼성카드",
    "현대캐피탈(직)", "MG캐피탈", "레드캡투어", "롯데렌탈", "현대캐피탈(플랫폼)",
]
# 위탁매출 거래처를 담은 컬럼 (필요시 변경)
_CONSIGNMENT_PARTNER_COLUMN = "분류3"


def _num(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)


def _compute_accounting_tables(master_df):
    """최종원가마스터에서 회계처리 표 3개의 값을 월별로 계산.

    반환: {"auto": df, "sales": df, "consign": df} (각 멀티헤더 표)
    """
    df = master_df.copy()
    df.columns = [str(c) for c in df.columns]

    month_series = (
        pd.to_numeric(df["회계월"], errors="coerce")
        if "회계월" in df.columns
        else pd.Series([pd.NA] * len(df), index=df.index)
    )
    sales_series = (
        df["매출구분"].astype(str).str.strip()
        if "매출구분" in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )

    def col(name):
        return _num(df[name]) if name in df.columns else pd.Series([0.0] * len(df), index=df.index)

    def partner_series():
        if _CONSIGNMENT_PARTNER_COLUMN in df.columns:
            return df[_CONSIGNMENT_PARTNER_COLUMN].astype(str).str.strip()
        return pd.Series([""] * len(df), index=df.index)

    auto_rows = [
        ("기초재고", 0), ("제조원가", 1),
        ("당기입고", 0), ("정상입고", 1), ("타계정입고", 1),
        ("당기출고", 0), ("정상출고", 1), ("자산출고", 1), ("기타출고", 1), ("기말재고", 0),
    ]
    auto_table = _build_accounting_table(auto_rows)

    sales_rows = [
        ("보험매출", 0), ("정비매출", 0), ("검사매출", 0), ("위탁매출", 0),
    ] + [(p, 1) for p in _CONSIGNMENT_PARTNERS] + [("기타", 1)]
    sales_table = _build_accounting_table(sales_rows)

    consign_rows = [
        ("기초재고", 0), ("제조원가", 1),
        ("당기입고", 0), ("정상입고", 1),
        ("당기출고", 0), ("제조원가", 1), ("위탁판매", 2), ("위탁매입", 2), ("위탁취소", 2),
        ("기말재고", 0),
    ]
    consign_table = _build_accounting_table(consign_rows)

    for month_label in _ACCOUNTING_MONTH_LABELS:
        if month_label == "누계":
            month_mask = pd.Series([True] * len(df), index=df.index)
        else:
            m = int(month_label.replace("월", ""))
            month_mask = month_series.eq(m)

        in_house = month_mask & sales_series.eq("사내매출")
        consign_m = month_mask & sales_series.eq("위탁매출")

        def s(mask, name):
            return float(col(name)[mask].sum())

        # ① 자동차 수불 (사내매출)
        base_q = s(in_house, "기초_수량"); base_a = s(in_house, "기초_금액")
        normal_in_q = s(in_house, "정상입고_수량"); normal_in_a = s(in_house, "정상입고_금액")
        transfer_in_q = s(in_house, "타처입고_수량"); transfer_in_a = s(in_house, "타처입고_금액")
        mfg_a = s(in_house, "제조원가_당월")
        in_q = normal_in_q + transfer_in_q
        in_a = normal_in_a + transfer_in_a
        normal_out_q = s(in_house, "정상출고_수량"); normal_out_a = s(in_house, "정상출고_금액")
        asset_out_q = s(in_house, "자산출고_수량"); asset_out_a = s(in_house, "자산출고_금액")
        etc_out_q = s(in_house, "기타출고_수량"); etc_out_a = s(in_house, "기타출고_금액")
        out_q = normal_out_q + asset_out_q + etc_out_q
        out_a = normal_out_a + asset_out_a + etc_out_a
        end_q = base_q + in_q - out_q
        end_a = base_a + in_a + mfg_a - out_a

        auto_values = {
            "기초재고": (base_q, base_a),
            "당기입고": (in_q, in_a),
            "정상입고": (normal_in_q, normal_in_a),
            "타계정입고": (transfer_in_q, transfer_in_a),
            "제조원가": (base_q, mfg_a),
            "당기출고": (out_q, out_a),
            "정상출고": (normal_out_q, normal_out_a),
            "자산출고": (asset_out_q, asset_out_a),
            "기타출고": (etc_out_q, etc_out_a),
            "기말재고": (end_q, end_a),
        }
        col_q = auto_table.columns.get_loc((month_label, "대수"))
        col_a = auto_table.columns.get_loc((month_label, "금액"))
        for row_pos, (name, level) in enumerate(auto_rows):
            q, a = auto_values[name]
            auto_table.iloc[row_pos, col_q] = q
            auto_table.iloc[row_pos, col_a] = a

        # ② 매출구분별
        def sales_count_amount(mask):
            return int(mask.sum()), float(col("제조원가")[mask].sum())

        ins_q, ins_a = sales_count_amount(month_mask & sales_series.eq("보험매출"))
        mnt_q, mnt_a = sales_count_amount(month_mask & sales_series.eq("정비매출"))
        insp_q, insp_a = sales_count_amount(month_mask & sales_series.eq("검사매출"))
        cons_q, cons_a = sales_count_amount(consign_m)

        sales_values = {
            "보험매출": (ins_q, ins_a),
            "정비매출": (mnt_q, mnt_a),
            "검사매출": (insp_q, insp_a),
            "위탁매출": (cons_q, cons_a),
        }
        partners = partner_series()
        partner_sum_q = 0
        partner_sum_a = 0.0
        for p in _CONSIGNMENT_PARTNERS:
            pq, pa = sales_count_amount(consign_m & partners.eq(p))
            sales_values[p] = (pq, pa)
            partner_sum_q += pq
            partner_sum_a += pa
        sales_values["기타"] = (cons_q - partner_sum_q, cons_a - partner_sum_a)

        col_q = sales_table.columns.get_loc((month_label, "대수"))
        col_a = sales_table.columns.get_loc((month_label, "금액"))
        for row_pos, (name, level) in enumerate(sales_rows):
            q, a = sales_values[name]
            sales_table.iloc[row_pos, col_q] = q
            sales_table.iloc[row_pos, col_a] = a

        # ③ 위탁 수불 (위탁매출)
        c_base_q = s(consign_m, "기초_수량")
        c_base_a = float(col("기초_금액")[consign_m & col("기초_수량").eq(1)].sum())
        c_base_mfg_a = float(col("제조원가")[consign_m & col("기초_수량").gt(0)].sum())
        c_in_q = s(consign_m, "정상입고_수량")
        c_in_a = float(col("제조원가")[consign_m & col("정상입고_수량").gt(0)].sum())
        c_out_q = s(consign_m, "정상출고_수량")
        c_out_a = s(consign_m, "정상출고_금액")
        status = (
            df["위탁출고구분"].astype(str).str.strip()
            if "위탁출고구분" in df.columns
            else pd.Series([""] * len(df), index=df.index)
        )
        def consign_status(value):
            mask = consign_m & status.eq(value)
            return s(mask, "정상출고_수량"), s(mask, "정상출고_금액")
        sale_q, sale_a = consign_status("위탁판매")
        buy_q, buy_a = consign_status("위탁매입")
        cancel_q, cancel_a = consign_status("위탁취소")
        c_end_q = c_base_q + c_in_q - c_out_q
        c_end_a = float(col("기말_금액")[consign_m & col("기말_수량").eq(1)].sum())

        consign_values_ordered = [
            (c_base_q, c_base_a),
            (c_base_q, c_base_mfg_a),
            (c_in_q, c_in_a),
            (c_in_q, c_in_a),
            (c_out_q, c_out_a),
            (c_out_q, c_out_a),
            (sale_q, sale_a),
            (buy_q, buy_a),
            (cancel_q, cancel_a),
            (c_end_q, c_end_a),
        ]
        col_q = consign_table.columns.get_loc((month_label, "대수"))
        col_a = consign_table.columns.get_loc((month_label, "금액"))
        for row_pos, value in enumerate(consign_values_ordered):
            q, a = value
            consign_table.iloc[row_pos, col_q] = q
            consign_table.iloc[row_pos, col_a] = a

    return {"auto": auto_table, "sales": sales_table, "consign": consign_table}


def _build_accounting_table(row_specs):
    """회계처리용 빈 표 생성.

    row_specs: [(행이름, 들여쓰기레벨), ...]
    반환: 멀티헤더(월/누계 × 대수/금액) DataFrame, 값은 0.
    중복 라벨은 zero-width space 로 고유화 (Styler 가 비고유 인덱스를 거부하므로).
    """
    columns = pd.MultiIndex.from_product(
        [_ACCOUNTING_MONTH_LABELS, ["대수", "금액"]]
    )
    index_labels = []
    seen = {}
    for name, level in row_specs:
        label = ("\u00a0" * (level * 4)) + name
        # 중복이면 보이지 않는 zero-width space(\u200b) 를 개수만큼 붙여 고유화
        count = seen.get(label, 0)
        seen[label] = count + 1
        if count > 0:
            label = label + ("\u200b" * count)
        index_labels.append(label)
    data = [[0] * len(columns) for _ in row_specs]
    df = pd.DataFrame(data, index=index_labels, columns=columns)
    return df


def _render_accounting_table(title, table, highlight_rows):
    """회계처리 표 하나 렌더링 (강조행 색상+볼드, 누계 컬럼 강조)."""
    st.markdown(f"**{title}**")

    highlight_positions = set(highlight_rows)
    n_rows = len(table)
    n_cols = len(table.columns)

    # 위치 기반 스타일 매트릭스 (중복 인덱스 라벨 대응)
    style_df = pd.DataFrame("", index=table.index, columns=table.columns)
    for r in range(n_rows):
        row_style = ""
        if r in highlight_positions:
            row_style = "background-color: #fce4d6; font-weight: bold;"
        for c in range(n_cols):
            col_tuple = table.columns[c]
            border = ""
            if isinstance(col_tuple, tuple) and col_tuple[0] == "누계":
                border = "border-left: 2px solid #c00000;"
            style_df.iloc[r, c] = row_style + border

    styler = table.style.format("{:,.0f}").apply(lambda _: style_df, axis=None)
    row_height_px = 35
    header_height_px = 65  # 멀티헤더 2줄 + 여유
    total_height = header_height_px + row_height_px * len(table) + 10  # 하단 여유
    # 표끼리(상품(자동차) 수불 / 상품(위탁) 수불) 라벨 폭이 달라 월 컬럼 경계가 어긋나지 않도록 고정
    st.dataframe(
        styler, use_container_width=True, height=total_height,
        column_config={"_index": st.column_config.Column(width=170)},
    )


def render_accounting_section(master_df=None):
    """회계처리 섹션: 상품(자동차) 수불 / 매출구분별 / 상품(위탁) 수불."""
    st.subheader("회계처리")

    if master_df is None or master_df.empty:
        st.info("최종 원가 마스터가 저장되면 회계처리 표가 채워집니다.")
        return

    selected_years_key = None
    if "회계연도" in master_df.columns:
        available_years = sorted(
            pd.to_numeric(master_df["회계연도"], errors="coerce").dropna().astype(int).unique().tolist(),
            reverse=True,
        )
        if available_years:
            _year_col, _ = st.columns([1, 3])
            with _year_col:
                selected_years = st.multiselect(
                    "회계연도", available_years,
                    default=[available_years[0]], key="accounting_section_years",
                )
            if not selected_years:
                st.info("회계연도를 선택하세요.")
                return
            master_df = master_df[
                pd.to_numeric(master_df["회계연도"], errors="coerce").isin(selected_years)
            ]
            if master_df.empty:
                st.info("선택한 연도의 데이터가 없습니다.")
                return
            selected_years_key = tuple(sorted(selected_years))

    # 마스터 파일이 안 바뀌고 연도 선택도 그대로면(경로+mtime+선택연도로 식별) 재계산하지 않도록 캐싱
    # (id(master_df)는 st.cache_data 를 거치며 매 리런마다 달라져 캐시 키로 못 씀)
    tables = _memoize(
        "accounting_tables", (_master_file_state(), selected_years_key),
        lambda: _compute_accounting_tables(master_df),
    )
    # 강조행: 소계/합계 (행 위치 기준)
    # ① 자동차: 기초재고0, 당기입고2, 당기출고5, 기말재고9
    _render_accounting_table("상품(자동차) 수불", tables["auto"], [0, 2, 5, 9])
    # ② 위탁: 기초재고0, 당기입고2, 당기출고4, 기말재고9
    _render_accounting_table("상품(위탁) 수불", tables["consign"], [0, 2, 4, 9])
    # # ③ 매출구분별: 위탁매출(3)만 강조
    # _render_accounting_table("매출구분별", tables["sales"], [3])     # 매출구분별 표


tab1, tab2 = st.tabs(["VIEW", "UPLOAD"])

with tab1:
    master_df, master_saved_at = load_final_master()

    if master_df is not None:
        col_space, col_btn = st.columns([8, 2])
        with col_btn:
            if st.button(
                "🗑️ 전체 데이터 초기화", type="primary", use_container_width=True,
                key="cost_reset_button",
            ):
                st.session_state['cost_delete_confirm'] = True

            st.markdown(
                f"<p style='text-align: right; color: gray; font-size: 0.75rem; margin-top: -10px;'>* 최근 업데이트: {master_saved_at or '-'}</p>",
                unsafe_allow_html=True
            )

        if st.session_state.get('cost_delete_confirm'):
            c1, c2 = st.columns(2)
            with c1:
                if st.button("✅ 삭제", use_container_width=True, key="cost_reset_confirm"):
                    delete_final_master()
                    st.session_state['cost_delete_confirm'] = False
                    st.rerun()
            with c2:
                if st.button("❌ 취소", use_container_width=True, key="cost_reset_cancel"):
                    st.session_state['cost_delete_confirm'] = False
                    st.rerun()

    # 회계처리 (먼저)
    render_accounting_section(master_df)

    st.divider()

    # 차량별 원가 (제목 옆 오른쪽에 다운로드 버튼)
    # 회계연도/회계월별로: 화면은 탭 분리, 엑셀 다운로드는 시트 분리
    col_title, col_dl = st.columns([8, 1], vertical_alignment="bottom")
    with col_title:
        st.subheader("차량별 원가")

    if master_df is not None:
        has_period = "회계연도" in master_df.columns and "회계월" in master_df.columns
        if has_period:
            ym_pairs = sorted({
                (int(y), int(m))
                for y, m in zip(
                    pd.to_numeric(master_df["회계연도"], errors="coerce").dropna(),
                    pd.to_numeric(master_df["회계월"], errors="coerce").dropna(),
                )
            })
        else:
            ym_pairs = []

        # 다운로드 엑셀: 회계월마다 시트 분리 (없으면 단일 시트)
        # + 컬럼 그룹별 배경색 + 헤더 볼드/큰글씨 + 천 단위 콤마
        def _build_monthly_split_excel_bytes(df):
            from io import BytesIO
            import datetime as _dt
            from openpyxl.utils import get_column_letter
            from cost_summary_preprocess import _apply_excel_column_group_styles

            _YEAR_KEYWORDS = ("연도", "년도")

            def _is_date_like(value):
                return isinstance(value, (_dt.datetime, _dt.date)) and not isinstance(value, bool)

            def _write_and_format_sheet(writer, sheet_name, sheet_df):
                sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]
                # 시트별 컬럼 그룹 색칠 + 헤더 스타일
                _apply_excel_column_group_styles(ws, sheet_df)

                # 숫자/날짜 컬럼 포맷 적용 (컬럼 단위로 한 번에 — 셀 단위 반복 대신)
                #  - 컬럼명에 '연도'/'년도' 포함: 포맷 적용 안 함 (천 단위 콤마도 X)
                #  - 숫자: #,##0
                #  - 날짜/datetime: 무조건 yyyy-mm-dd (시간 부분 제거)
                for col_idx, col_name in enumerate(sheet_df.columns, start=1):
                    if col_name is None or any(kw in str(col_name) for kw in _YEAR_KEYWORDS):
                        continue

                    col_series = sheet_df.iloc[:, col_idx - 1]
                    if pd.api.types.is_bool_dtype(col_series):
                        continue

                    letter = get_column_letter(col_idx)
                    if pd.api.types.is_datetime64_any_dtype(col_series):
                        is_date_col = True
                    elif pd.api.types.is_numeric_dtype(col_series):
                        is_date_col = False
                    else:
                        # dtype 만으로 판단 안 되는 object 컬럼만 값 검사 (드문 케이스)
                        is_date_col = col_series.apply(_is_date_like).any()

                    if is_date_col:
                        # datetime 값만 시간 부분 제거 (date 값은 그대로 유지, 원본 동일)
                        for row_idx, val in enumerate(col_series, start=2):
                            if isinstance(val, _dt.datetime):
                                ws.cell(row=row_idx, column=col_idx).value = val.date()
                        ws.column_dimensions[letter].number_format = "yyyy-mm-dd"
                    elif pd.api.types.is_numeric_dtype(col_series) or col_series.apply(
                        lambda v: isinstance(v, (int, float)) and not isinstance(v, bool)
                    ).any():
                        ws.column_dimensions[letter].number_format = "#,##0"

            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                if ym_pairs:
                    yr_series = pd.to_numeric(df["회계연도"], errors="coerce")
                    mo_series = pd.to_numeric(df["회계월"], errors="coerce")
                    for y, m in ym_pairs:
                        mask = yr_series.eq(y) & mo_series.eq(m)
                        sub = df[mask]
                        if sub.empty:
                            continue
                        _write_and_format_sheet(writer, f"{y}-{m:02d}", sub)
                else:
                    _write_and_format_sheet(writer, "최종원가마스터", df)
            return buf.getvalue()

        with col_dl:
            st.download_button(
                "다운로드",
                # 마스터 파일이 안 바뀌면(경로+mtime로 식별) 전체 워크북을 다시 만들지 않도록 캐싱.
                # master_df 자체는 st.cache_data 를 거치며 매 호출마다 deep-copy 되어 id()가 매번
                # 달라지므로, id(master_df) 대신 파일의 (경로, mtime)을 지문으로 사용해야 실제로 캐싱됨.
                data=_memoize(
                    "view_master_download", _master_file_state() or (None, None),
                    lambda: _build_monthly_split_excel_bytes(master_df),
                ),
                file_name=f"cost_summary_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="master_download_top",
                use_container_width=True,
            )

    if master_df is None:
        st.info("저장된 최종 원가 마스터가 없습니다. UPLOAD 탭에서 생성 후 '최종 마스터 저장'을 눌러주세요.")
    else:
        # 화면: 회계월별 탭 분리 (오름차순)
        if ym_pairs:
            tab_labels = [f"{y}-{m:02d}" for y, m in ym_pairs]
            monthly_tabs = st.tabs(tab_labels)
            yr_series = pd.to_numeric(master_df["회계연도"], errors="coerce")
            mo_series = pd.to_numeric(master_df["회계월"], errors="coerce")
            for (y, m), tab_obj in zip(ym_pairs, monthly_tabs):
                with tab_obj:
                    mask = yr_series.eq(y) & mo_series.eq(m)
                    sub = master_df[mask]
                    st.caption(f"{y}-{m:02d} · {len(sub):,}건")
                    st.dataframe(dataframe_for_display(sub), use_container_width=True)
        else:
            st.dataframe(dataframe_for_display(master_df), use_container_width=True)

with tab2:
    settlement_year, settlement_month = render_settlement_selector()
    dfs, product_id_df, base_sheets = render_base_upload(settlement_year, settlement_month)

    st.divider()
    render_section_heading("매출원가")
    purchase_cost_sheet_dfs, purchase_cost_sheets_raw = render_purchase_cost_upload(
        product_id_df, dfs, settlement_year, settlement_month,
    )

    st.divider()
    manufacturing_cost_sheet_dfs, manufacturing_sheets_raw = render_manufacturing_cost_upload(
        product_id_df, settlement_year, settlement_month,
    )

    st.divider()
    cost_driver_dfs, cost_driver_sheets_raw = render_cost_driver_upload(settlement_year, settlement_month)
    combined_cost_driver_df = render_combined_cost_driver(
        cost_driver_dfs, settlement_year, settlement_month, product_id_df,
    )

    st.divider()
    verification_sheets = render_verification_sheet_upload()

    st.divider()
    final_cost_df = render_final_cost(
        product_id_df, purchase_cost_sheet_dfs, dfs, settlement_month,
        manufacturing_cost_sheet_dfs=manufacturing_cost_sheet_dfs,
        verification_sheets=verification_sheets,
        settlement_year=settlement_year,
        cost_driver_dfs=cost_driver_dfs,
        combined_cost_driver_df=combined_cost_driver_df,
    )

    # 최종 마스터 저장 — 선택한 기준연/월부터, 업로드된 데이터에 있는 마지막 달까지 전부 처리.
    # (중간 달이 바뀌면 그 뒤 달도 전부 다시 계산해야 하므로, 선택한 달 하나만이 아니라
    # 그 이후 달을 전부 순서대로 다시 계산 후 저장한다 — 기초재고는 바로 앞 달 계산 결과를
    # 자동으로 이어받는다.)
    st.divider()
    if final_cost_df is not None and not final_cost_df.empty:
        if st.button("기준월부터 재계산", key="save_final_master", type="primary"):
            available_periods = detect_available_periods(base_sheets)
            periods_from_selected = [
                p for p in available_periods if p >= (settlement_year, settlement_month)
            ] or [(settlement_year, settlement_month)]
            with st.spinner(f"{len(periods_from_selected)}개월 처리 중..."):
                batch_results = run_batch_periods(
                    base_sheets, purchase_cost_sheets_raw, manufacturing_sheets_raw,
                    cost_driver_sheets_raw, verification_sheets, periods_from_selected,
                )
            st.dataframe(pd.DataFrame(batch_results), use_container_width=True, hide_index=True)
            if any(str(r["상태"]).startswith("오류") for r in batch_results):
                st.warning("일부 기간에서 오류가 발생했습니다. 위 표를 확인하세요.")
            else:
                st.success(
                    f"{settlement_year}년 {settlement_month}월부터 {len(periods_from_selected)}개월치 "
                    "최종 마스터가 저장되었습니다. VIEW 탭에서 확인하세요."
                )
    else:
        st.caption("최종 원가 데이터가 생성되면 '기준월부터 재계산' 버튼이 활성화됩니다.")